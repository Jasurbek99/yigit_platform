import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

wb = openpyxl.load_workbook('d:/projects/yigit_platform/data/Export_contracts_2025-2026.xlsx', read_only=True, data_only=True)
ws = wb['2-Sales']

# Print header and first 10 data rows
print("=== HEADER ===")
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    for i, v in enumerate(row):
        print(f"  Col {i} ({chr(65+i)}): {v!r}")

print("\n=== FIRST 15 DATA ROWS (raw) ===")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=16, values_only=True), start=2):
    print(f"Row {i}: {list(row[:8])}")

wb.close()
