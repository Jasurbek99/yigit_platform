"""Import invoices (and one-time contracts) from the 2-Sales sheet.

Reads the 2-Sales sheet of Export_contracts_2025-2026.xlsx and writes:
  - apps.contracts.Contract  (one-time / Pattern B contracts not in DB)
  - apps.contracts.Invoice   (both Pattern A and Pattern B)

Pattern classification:
  Pattern A -- col G (serial_no_of_truck) is populated and non-zero.
               Row represents a truck dispatch against a multi-truck master contract
               from sheet 1-Contracts.
  Pattern B -- col G is NULL / zero / '-'.
               Row represents a one-time ad-hoc shipment; the contract string in
               col D is treated as the contract_number for a new ONE_TIME contract.

Summary-leak rows (col J quantity > 50 000 AND col G serial not > 0) are skipped.
Completely empty rows (only the sequence number in col A) are also skipped.

Usage:
    python manage.py import_invoices_from_2sales --dry-run
    python manage.py import_invoices_from_2sales --commit
    python manage.py import_invoices_from_2sales --commit --file /abs/path
    python manage.py import_invoices_from_2sales --dry-run --max-rows 100

Flags:
    --dry-run        (default) Classify, report counts, write skipped CSV -- NO DB writes.
    --commit         Real writes wrapped in transaction.atomic().
    --file PATH      Override default Excel location.
    --max-rows N     Stop after N data rows (testing aid).

Outputs:
    data/import_invoices_skipped.csv  -- skipped-row audit trail (written in both modes).
"""
from __future__ import annotations

import csv
import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.contracts.models import Contract, Invoice
from apps.contracts.services.rollup import rollup_contract_totals
from apps.core.models import ExportFirm, ImportFirm, Season


BATCH_SIZE = 500

# ── Column indices (0-based) ───────────────────────────────────────────────────
# Header row: 0=seq, 1=seller, 2=buyer, 3=contract, 4=invoice_date,
#             5=total_trucks, 6=serial_no_of_truck, 7=inv_no,
#             8=incoterm, 9=quantity_kg, 10=usd, 11=truck_plate,
#             12=passport_sdelka, 13=scan, 14=r15_note
C_SEQ = 0
C_SELLER = 1
C_BUYER = 2
C_CONTRACT = 3
C_INV_DATE = 4
C_TOTAL_TRUCKS = 5
C_SERIAL = 6
C_INV_NO = 7
C_INCOTERM = 8
C_QTY_KG = 9
C_USD = 10
C_TRUCK_PLATE = 11
C_PASSPORT = 12
C_SCAN = 13
C_R15 = 14


# ── helpers ────────────────────────────────────────────────────────────────────

def _clean(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _safe_decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _safe_int(value) -> Optional[int]:
    if value is None:
        return None
    try:
        v = int(str(value).strip())
        return v
    except (TypeError, ValueError):
        return None


def _parse_date(value) -> Optional[datetime.date]:
    """Convert Excel cell value to date. Handles datetime objects and strings."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _is_serial_populated(value) -> bool:
    """Return True when serial_no_of_truck looks like a real serial number (>=1)."""
    if value is None:
        return False
    s = str(value).strip()
    if s in ('', '-', '0', 'None'):
        return False
    try:
        return int(float(s)) >= 1
    except (ValueError, TypeError):
        return False


class Command(BaseCommand):
    help = 'Import invoices (and one-time contracts) from 2-Sales sheet.'

    def add_arguments(self, parser):
        default_path = (
            Path(settings.BASE_DIR).parent / 'data' / 'Export_contracts_2025-2026.xlsx'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Classify and report counts -- no DB writes (CSV is still written).',
        )
        parser.add_argument(
            '--commit',
            action='store_true',
            default=False,
            help='Execute real writes.',
        )
        parser.add_argument(
            '--file',
            type=Path,
            default=default_path,
            metavar='PATH',
            help=f'Excel workbook path (default: {default_path})',
        )
        parser.add_argument(
            '--max-rows',
            type=int,
            default=None,
            metavar='N',
            help='Cap at N data rows (testing aid).',
        )

    def handle(self, *args, **opts):
        path: Path = opts['file']
        dry_run: bool = opts['dry_run'] or not opts['commit']
        max_rows: Optional[int] = opts['max_rows']

        if not path.exists():
            raise CommandError(f'File not found: {path}')

        # CSV output lives alongside the Excel file
        csv_path = path.parent / 'import_invoices_skipped.csv'

        self.stdout.write(f'Loading: {path}')
        wb = load_workbook(path, data_only=True, read_only=True)

        if '2-Sales' not in wb.sheetnames:
            raise CommandError(
                f"Sheet '2-Sales' not found. Available: {wb.sheetnames}"
            )

        ws = wb['2-Sales']

        # ── Pre-load reference data ────────────────────────────────────────────
        export_firms: dict[str, ExportFirm] = {
            f.code: f for f in ExportFirm.objects.all()
        }
        import_firms_by_short: dict[str, ImportFirm] = {
            str(f.name_short or '').strip().lower(): f
            for f in ImportFirm.objects.all()
            if f.name_short
        }
        import_firms_by_company: dict[str, ImportFirm] = {
            str(f.name_company or '').strip().lower(): f
            for f in ImportFirm.objects.all()
            if f.name_company
        }
        season: Optional[Season] = (
            Season.objects.filter(is_active=True).first()
            or Season.objects.order_by('-start_date').first()
        )
        if season is None:
            raise CommandError('No Season in DB. Run seed_data first.')

        # Existing contracts (for Pattern A matching AND Pattern B collision check)
        # Key = contract_number.lower()
        existing_contracts: dict[str, Contract] = {
            c.contract_number.lower(): c
            for c in Contract.objects.all()
        }

        # Existing invoices: (contract_id, invoice_number) to avoid unique_together collision
        existing_invoice_keys: set[tuple[int, int]] = set(
            Invoice.objects.values_list('contract_id', 'invoice_number')
        )

        # ── Counters / collectors ──────────────────────────────────────────────
        cnt_rows_total = 0
        cnt_empty = 0
        cnt_summary_leaks = 0
        cnt_pat_a = 0
        cnt_pat_b = 0

        cnt_a_matched = 0
        cnt_a_no_contract = 0
        cnt_a_dq = 0

        cnt_b_unique_new = 0      # distinct contract strings we'll create
        cnt_b_invoices = 0
        cnt_b_skipped_fk = 0
        cnt_b_skipped_collision = 0  # matched existing contract via collision check

        skipped_rows: list[dict] = []  # audit CSV

        # Pattern B accumulators
        # new_one_time_contracts: contract_number -> Contract (unsaved)
        new_one_time_contracts: dict[str, Contract] = {}
        # pat_b_invoice_payloads: list of dicts, will materialise after bulk insert
        pat_b_invoice_payloads: list[dict] = []

        # Pattern A invoice objects
        pat_a_invoices: list[Invoice] = []
        affected_contract_ids: set[int] = set()

        # ── Row processing ─────────────────────────────────────────────────────
        data_rows_seen = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[C_SEQ] is None:
                continue

            # Row-count cap (testing only)
            if max_rows is not None and data_rows_seen >= max_rows:
                break
            data_rows_seen += 1

            cnt_rows_total += 1
            excel_row_num = data_rows_seen + 1  # approximate; 1-based with header

            # 1. Skip completely empty rows (only seq no populated)
            non_none = [v for v in row if v is not None]
            if len(non_none) <= 1:
                cnt_empty += 1
                continue

            # 2. Summary-leak detection: qty > 50,000 and not a real serial
            raw_qty = row[C_QTY_KG] if len(row) > C_QTY_KG else None
            qty = _safe_decimal(raw_qty)
            serial_raw = row[C_SERIAL] if len(row) > C_SERIAL else None
            if qty is not None and qty > Decimal('50000') and not _is_serial_populated(serial_raw):
                cnt_summary_leaks += 1
                skipped_rows.append({
                    'excel_row_num': excel_row_num,
                    'reason': 'summary_leak_qty>50000',
                    'contract_string': _clean(row[C_CONTRACT] if len(row) > C_CONTRACT else None),
                    'seller': _clean(row[C_SELLER] if len(row) > C_SELLER else None),
                    'buyer': _clean(row[C_BUYER] if len(row) > C_BUYER else None),
                    'invoice_date': _clean(row[C_INV_DATE] if len(row) > C_INV_DATE else None),
                    'quantity_kg': str(qty),
                })
                continue

            # 3. Classify
            is_pat_a = _is_serial_populated(serial_raw)
            if is_pat_a:
                cnt_pat_a += 1
            else:
                cnt_pat_b += 1

            # ── Common field extraction ────────────────────────────────────────
            seller_raw = _clean(row[C_SELLER] if len(row) > C_SELLER else None)
            buyer_raw = _clean(row[C_BUYER] if len(row) > C_BUYER else None)
            contract_str = _clean(row[C_CONTRACT] if len(row) > C_CONTRACT else None)
            inv_date = _parse_date(row[C_INV_DATE] if len(row) > C_INV_DATE else None)
            inv_no_raw = row[C_INV_NO] if len(row) > C_INV_NO else None
            incoterm = _clean(row[C_INCOTERM] if len(row) > C_INCOTERM else None)[:10]
            total_usd = _safe_decimal(row[C_USD] if len(row) > C_USD else None)
            passport = _clean(row[C_PASSPORT] if len(row) > C_PASSPORT else None)[:100]
            scan_raw = row[C_SCAN] if len(row) > C_SCAN else None
            scan_uploaded = bool(scan_raw) and str(scan_raw).strip() not in (
                'False', '0', '-', 'None', 'yatyryldy', 'iptal', 'YZA SUYSIRILDI', ''
            )
            serial_int: Optional[int] = _safe_int(serial_raw) if is_pat_a else None

            # Compute price_per_kg (may be None if qty or usd is None/zero)
            price_per_kg: Optional[Decimal] = None
            if qty and total_usd and qty > 0:
                try:
                    price_per_kg = (total_usd / qty).quantize(Decimal('0.0001'))
                except Exception:
                    price_per_kg = None

            # Coerce invoice_number
            inv_no: Optional[int] = None
            try:
                inv_no_str = str(inv_no_raw).strip() if inv_no_raw is not None else ''
                if inv_no_str:
                    inv_no = int(float(inv_no_str))
            except (ValueError, TypeError):
                inv_no = None

            # ── Resolve firms for FK fields on Invoice ─────────────────────────
            export_firm = export_firms.get(seller_raw)
            buyer_lower = buyer_raw.lower()
            import_firm = (
                import_firms_by_short.get(buyer_lower)
                or import_firms_by_company.get(buyer_lower)
            )

            # ================================================================
            # PATTERN A processing
            # ================================================================
            if is_pat_a:
                if not contract_str:
                    cnt_a_no_contract += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_a_no_contract_string',
                        'contract_string': '',
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': str(inv_date),
                        'quantity_kg': str(qty),
                    })
                    continue

                # Look up contract — exact then case-insensitive
                contract_obj = existing_contracts.get(contract_str.lower())
                if contract_obj is None:
                    cnt_a_no_contract += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_a_contract_not_in_db',
                        'contract_string': contract_str,
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': str(inv_date),
                        'quantity_kg': str(qty),
                    })
                    continue

                # Validate invoice_number
                if inv_no is None:
                    cnt_a_dq += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_a_invoice_number_non_integer',
                        'contract_string': contract_str,
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': str(inv_date),
                        'quantity_kg': str(qty),
                    })
                    continue

                # Check invoice_date
                if inv_date is None:
                    cnt_a_dq += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_a_missing_invoice_date',
                        'contract_string': contract_str,
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': '',
                        'quantity_kg': str(qty),
                    })
                    continue

                # Check unique_together collision
                key = (contract_obj.id, inv_no)
                if key in existing_invoice_keys:
                    cnt_a_dq += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_a_invoice_already_exists',
                        'contract_string': contract_str,
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': str(inv_date),
                        'quantity_kg': str(qty),
                    })
                    continue
                existing_invoice_keys.add(key)

                cnt_a_matched += 1
                affected_contract_ids.add(contract_obj.id)
                pat_a_invoices.append(Invoice(
                    contract=contract_obj,
                    invoice_number=inv_no,
                    invoice_date=inv_date,
                    serial_truck_number=serial_int,
                    export_firm=export_firm,
                    import_firm=import_firm,
                    incoterm=incoterm,
                    quantity_kg=qty,
                    price_per_kg=price_per_kg,
                    total_usd=total_usd,
                    passport_sdelka=passport,
                    scan_uploaded=scan_uploaded,
                    status=Invoice.STATUS_SENT,
                ))

            # ================================================================
            # PATTERN B processing
            # ================================================================
            else:
                if not buyer_raw or import_firm is None:
                    cnt_b_skipped_fk += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_b_import_firm_not_found',
                        'contract_string': contract_str,
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': str(inv_date),
                        'quantity_kg': str(qty),
                    })
                    continue

                if not seller_raw or export_firm is None:
                    cnt_b_skipped_fk += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_b_export_firm_not_found',
                        'contract_string': contract_str,
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': str(inv_date),
                        'quantity_kg': str(qty),
                    })
                    continue

                if inv_date is None:
                    cnt_b_skipped_fk += 1
                    skipped_rows.append({
                        'excel_row_num': excel_row_num,
                        'reason': 'pat_b_missing_invoice_date',
                        'contract_string': contract_str,
                        'seller': seller_raw,
                        'buyer': buyer_raw,
                        'invoice_date': '',
                        'quantity_kg': str(qty),
                    })
                    continue

                # Determine the contract key for this row.
                # If contract_str is empty, synthesise one from seller+buyer+date
                # so we can still group per "deal".
                if contract_str:
                    contract_key = contract_str
                else:
                    contract_key = f'ONETIME/{seller_raw}/{buyer_raw}/{inv_date}'

                # Check if this contract_key already exists in the DB
                # (handles the 4 Pattern B rows that collide with master contracts)
                contract_obj = existing_contracts.get(contract_key.lower())
                if contract_obj is not None:
                    # Already in DB — attach invoice to existing contract
                    cnt_b_skipped_collision += 1
                    # Reuse Pattern A logic: wire to existing contract
                    inv_no_use = inv_no if inv_no is not None else 1
                    key = (contract_obj.id, inv_no_use)
                    if key in existing_invoice_keys:
                        # True duplicate — skip
                        skipped_rows.append({
                            'excel_row_num': excel_row_num,
                            'reason': 'pat_b_collision_invoice_already_exists',
                            'contract_string': contract_key,
                            'seller': seller_raw,
                            'buyer': buyer_raw,
                            'invoice_date': str(inv_date),
                            'quantity_kg': str(qty),
                        })
                        continue
                    existing_invoice_keys.add(key)
                    affected_contract_ids.add(contract_obj.id)
                    cnt_b_invoices += 1
                    pat_b_invoice_payloads.append({
                        '_contract_key': contract_key,
                        '_use_existing_id': contract_obj.id,
                        'invoice_number': inv_no_use,
                        'invoice_date': inv_date,
                        'serial_truck_number': None,
                        'export_firm': export_firm,
                        'import_firm': import_firm,
                        'incoterm': incoterm,
                        'quantity_kg': qty,
                        'price_per_kg': price_per_kg,
                        'total_usd': total_usd,
                        'passport_sdelka': passport,
                        'scan_uploaded': scan_uploaded,
                        'status': Invoice.STATUS_SENT,
                    })
                    continue

                # New one-time contract
                if contract_key not in new_one_time_contracts:
                    cnt_b_unique_new += 1
                    new_one_time_contracts[contract_key] = Contract(
                        contract_number=contract_key,
                        season=season,
                        export_firm=export_firm,
                        import_firm=import_firm,
                        contract_type='ONE_TIME',
                        incoterm=incoterm,
                        start_date=inv_date,
                        planned_trucks=1,
                        planned_quantity_kg=qty,
                        planned_amount_usd=total_usd,
                        status=Contract.STATUS_ACTIVE,
                    )

                cnt_b_invoices += 1
                # Assign invoice_number = 1 (only one invoice per ONE_TIME contract)
                # If the same contract_key appears a second time, give it 2, etc.
                # Track count per contract_key to auto-increment invoice_number
                pat_b_invoice_payloads.append({
                    '_contract_key': contract_key,
                    '_use_existing_id': None,  # will be set after bulk_create
                    'invoice_number': None,      # will be computed after bulk_create
                    'invoice_date': inv_date,
                    'serial_truck_number': None,
                    'export_firm': export_firm,
                    'import_firm': import_firm,
                    'incoterm': incoterm,
                    'quantity_kg': qty,
                    'price_per_kg': price_per_kg,
                    'total_usd': total_usd,
                    'passport_sdelka': passport,
                    'scan_uploaded': scan_uploaded,
                    'status': Invoice.STATUS_SENT,
                })

        # ── Post-loop: assign invoice_number for each Pattern B payload ────────
        # Invoice_number is sequential per contract_key (1-based)
        contract_key_inv_counter: dict[str, int] = {}
        for payload in pat_b_invoice_payloads:
            if payload['_use_existing_id'] is not None:
                continue  # invoice_number already set for collision rows
            ck = payload['_contract_key']
            contract_key_inv_counter[ck] = contract_key_inv_counter.get(ck, 0) + 1
            payload['invoice_number'] = contract_key_inv_counter[ck]

        # ── Write skipped-rows CSV (always, both modes) ────────────────────────
        with open(csv_path, 'w', newline='', encoding='utf-8') as fh:
            writer = csv.DictWriter(fh, fieldnames=[
                'excel_row_num', 'reason', 'contract_string',
                'seller', 'buyer', 'invoice_date', 'quantity_kg',
            ])
            writer.writeheader()
            writer.writerows(skipped_rows)
        self.stdout.write(f'Skipped-rows CSV: {csv_path} ({len(skipped_rows)} rows)')

        # ── DB writes (only when --commit) ─────────────────────────────────────
        if not dry_run:
            self._execute_writes(
                new_one_time_contracts=new_one_time_contracts,
                pat_a_invoices=pat_a_invoices,
                pat_b_invoice_payloads=pat_b_invoice_payloads,
                affected_contract_ids=affected_contract_ids,
            )

        # ── Report ─────────────────────────────────────────────────────────────
        total_data_rows = cnt_rows_total - cnt_empty - cnt_summary_leaks
        self.stdout.write('')
        self.stdout.write('=== Pattern classification ===')
        self.stdout.write(f'Pattern A rows (multi-truck):              {cnt_pat_a:,}')
        self.stdout.write(f'Pattern B rows (one-time):                 {cnt_pat_b:,}')
        self.stdout.write(f'Summary-leak rows skipped:                 {cnt_summary_leaks:,}')
        self.stdout.write(f'Completely empty rows skipped:             {cnt_empty:,}')
        self.stdout.write(f'Total rows with seq no:                    {cnt_rows_total:,}')
        self.stdout.write('')
        self.stdout.write('=== Pattern A -- multi-truck invoices ===')
        self.stdout.write(f'Matched to existing contract:              {cnt_a_matched:,}')
        self.stdout.write(
            self.style.WARNING(
                f'Skipped (no contract match):               {cnt_a_no_contract:,} -> CSV'
            )
        )
        self.stdout.write(
            self.style.WARNING(
                f'Skipped (other DQ):                        {cnt_a_dq:,} -> CSV'
            )
        )
        self.stdout.write('')
        self.stdout.write('=== Pattern B -- one-time contracts ===')
        self.stdout.write(
            f'Unique new one-time contracts to create:   {cnt_b_unique_new:,}'
        )
        self.stdout.write(
            f'Pattern B rows attached to existing ctrs:  {cnt_b_skipped_collision:,}'
        )
        self.stdout.write(f'Pattern B invoices to create:              {cnt_b_invoices:,}')
        self.stdout.write(
            self.style.WARNING(
                f'Skipped (FK miss on seller/buyer/date):    {cnt_b_skipped_fk:,} -> CSV'
            )
        )
        self.stdout.write('')

        # IMPORTANT: flag the ~53% Pattern B FK miss rate so user sees it upfront
        if cnt_b_skipped_fk > 100:
            self.stdout.write(self.style.WARNING(
                f'WARNING: {cnt_b_skipped_fk:,} Pattern B rows skipped due to buyer not in DB. '
                f'Top missing buyers include: SAH FRUT, TransAsia Trade, IP Tursynbayew. '
                f'Run import_firms_from_excel with a buyer list first, then re-run this command.'
            ))
            self.stdout.write('')

        self.stdout.write('=== Final tally ===')
        self.stdout.write(
            self.style.SUCCESS(f'New Contract rows (one-time):              {cnt_b_unique_new:,}')
        )
        self.stdout.write(
            self.style.SUCCESS(
                f'New Invoice rows:                          '
                f'{cnt_a_matched + cnt_b_invoices:,}'
                f'  (A={cnt_a_matched:,} + B={cnt_b_invoices:,})'
            )
        )
        self.stdout.write(
            f'Affected contracts (rollup):               '
            f'{len(affected_contract_ids) + cnt_b_unique_new:,}'
        )

        if dry_run:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(
                '[DRY-RUN -- no writes performed. Re-run with --commit to write.]'
            ))
        else:
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS('Writes committed.'))

    def _execute_writes(
        self,
        new_one_time_contracts: dict[str, Contract],
        pat_a_invoices: list[Invoice],
        pat_b_invoice_payloads: list[dict],
        affected_contract_ids: set[int],
    ) -> None:
        """Bulk-create one-time contracts + invoices inside a single transaction.

        Rollup is called ONCE per affected contract AFTER all inserts, never in
        a loop during inserts.  bulk_create bypasses Invoice.save(), so the
        rollup that Invoice.save() would normally trigger does NOT fire — we
        call it manually here.
        """
        with transaction.atomic():
            # 1. Insert new one-time contracts
            new_contract_list = list(new_one_time_contracts.values())
            if new_contract_list:
                # bulk_create on MSSQL doesn't return PKs reliably,
                # so we re-fetch immediately after.
                Contract.objects.bulk_create(new_contract_list, batch_size=BATCH_SIZE)

            # 2. Re-fetch all contracts (including newly created) for ID lookup
            all_contracts_by_key: dict[str, Contract] = {
                c.contract_number.lower(): c
                for c in Contract.objects.all()
            }

            # 3. Build Pattern B Invoice objects
            pat_b_invoices: list[Invoice] = []
            for payload in pat_b_invoice_payloads:
                ck = payload['_contract_key']
                existing_id = payload['_use_existing_id']

                if existing_id is not None:
                    contract_obj = Contract(id=existing_id)
                else:
                    contract_obj = all_contracts_by_key.get(ck.lower())
                    if contract_obj is None:
                        self.stderr.write(
                            f'WARN: contract key {ck!r} not found after bulk_create -- skipping invoice'
                        )
                        continue

                affected_contract_ids.add(contract_obj.id)
                pat_b_invoices.append(Invoice(
                    contract_id=contract_obj.id,
                    invoice_number=payload['invoice_number'],
                    invoice_date=payload['invoice_date'],
                    serial_truck_number=payload['serial_truck_number'],
                    export_firm=payload['export_firm'],
                    import_firm=payload['import_firm'],
                    incoterm=payload['incoterm'],
                    quantity_kg=payload['quantity_kg'],
                    price_per_kg=payload['price_per_kg'],
                    total_usd=payload['total_usd'],
                    passport_sdelka=payload['passport_sdelka'],
                    scan_uploaded=payload['scan_uploaded'],
                    status=payload['status'],
                ))

            # 4. Bulk insert Pattern A invoices
            if pat_a_invoices:
                Invoice.objects.bulk_create(pat_a_invoices, batch_size=BATCH_SIZE)

            # 5. Bulk insert Pattern B invoices
            if pat_b_invoices:
                Invoice.objects.bulk_create(pat_b_invoices, batch_size=BATCH_SIZE)

            # 6. Rollup ONCE per affected contract — NEVER inside a row loop
            for cid in affected_contract_ids:
                rollup_contract_totals(cid)
