"""Import June-2026 sales from the 2-Sales sheet and bridge them to existing
shipments (ADR-023 Slice 3, narrowed scope).

June shipments already exist in the DB (imported elsewhere). This command does
NOT create shipments. For each June 2-Sales row (one export firm's share of a
truck) it:
  - matches the existing Shipment by (normalised truck_plate + date);
  - ensures a ShipmentFirmSplit for that export firm (creates from the row if
    missing);
  - resolves the Contract by the canonical (export_firm, year, seq) identity
    parsed from the row's contract string — links if it exists, else creates a
    one-time contract with that number;
  - creates/updates the ContractSale bridge keyed by (shipment, export_firm),
    carrying the row's invoice number/date, weight and amount.

Dry-run by default (no writes); pass --commit to persist. Idempotent: re-running
matches by the unique keys and updates rather than duplicating.

    python manage.py import_2sales_june_sales            # dry-run report
    python manage.py import_2sales_june_sales --commit   # persist
"""
from __future__ import annotations

import csv
import datetime
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.contracts.models import Contract, ContractSale
from apps.contracts.services.contract_number import parse_contract_number
from apps.core.models import ExportFirm, ImportFirm, Season
from apps.export.models import Shipment, ShipmentFirmSplit
from apps.contracts.management.commands.import_invoices_from_2sales import (
    _clean,
    _parse_date,
    _safe_decimal,
    _safe_int,
)

# Column indices (0-based) — mirror import_invoices_from_2sales.
C_SELLER, C_BUYER, C_CONTRACT, C_INV_DATE = 1, 2, 3, 4
C_INV_NO, C_QTY, C_USD, C_TRUCK, C_PASSPORT = 7, 9, 10, 11, 12

DEFAULT_FILE = 'data/Export_contracts_2025-2026.xlsx'
SHEET = '2-Sales'
BATCH = 500


def _norm_plate(value) -> str:
    """Normalise a truck plate for matching: upper, no spaces."""
    return ''.join(str(value or '').split()).upper()


class Command(BaseCommand):
    help = 'Import June-2026 2-Sales rows as ContractSale bridges to existing shipments.'

    def add_arguments(self, parser):
        parser.add_argument('--commit', action='store_true', help='Persist (default: dry-run).')
        parser.add_argument('--file', type=str, default=None, help='Override Excel path.')
        parser.add_argument('--year', type=int, default=2026)
        parser.add_argument('--month', type=int, default=6)

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        commit = opts['commit']
        year, month = opts['year'], opts['month']
        path = Path(opts['file']) if opts['file'] else Path(settings.BASE_DIR).parent / DEFAULT_FILE
        if not path.exists():
            path = Path(settings.BASE_DIR) / DEFAULT_FILE
        self.stdout.write(f'File: {path} | sheet {SHEET} | scope {year}-{month:02d} | '
                          f'{"COMMIT" if commit else "DRY-RUN"}')

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb[SHEET]

        # Reference maps.
        export_firms = {f.code: f for f in ExportFirm.objects.all()}
        imp_by_short = {str(f.name_short or '').strip().lower(): f
                        for f in ImportFirm.objects.all() if f.name_short}
        imp_by_company = {str(f.name_company or '').strip().lower(): f
                          for f in ImportFirm.objects.all()}
        # Shipment index by normalised plate → [(date, shipment)] — ALL months,
        # because the invoice date can lag the loading date by 1-2 weeks.
        ship_by_plate: dict[str, list[tuple[datetime.date, Shipment]]] = {}
        for s in (Shipment.objects.exclude(truck_plate__isnull=True)
                  .exclude(truck_plate='')):
            ship_by_plate.setdefault(_norm_plate(s.truck_plate), []).append((s.date, s))
        self.stdout.write(f'Indexed {len(ship_by_plate)} distinct plates across all shipments.')

        season = Season.objects.filter(is_active=True).order_by('-start_date').first()

        stats = dict(june=0, skipped_nonstd=0, skipped_no_firm=0, skipped_no_buyer=0,
                     skipped_unmatched=0, splits_created=0, contracts_linked=0,
                     contracts_created=0, sales_created=0, sales_updated=0)
        skipped: list[dict] = []
        actions: list[dict] = []  # deferred work for commit

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            d = _parse_date(row[C_INV_DATE] if len(row) > C_INV_DATE else None)
            if not (isinstance(d, datetime.date) and d.year == year and d.month == month):
                continue
            stats['june'] += 1

            seller = _clean(row[C_SELLER])
            buyer = _clean(row[C_BUYER])
            contract_str = _clean(row[C_CONTRACT])
            truck = _clean(row[C_TRUCK])
            inv_no = _safe_int(row[C_INV_NO] if len(row) > C_INV_NO else None)
            qty = _safe_decimal(row[C_QTY] if len(row) > C_QTY else None)
            usd = _safe_decimal(row[C_USD] if len(row) > C_USD else None)
            passport = _clean(row[C_PASSPORT] if len(row) > C_PASSPORT else None)

            def skip(reason):
                skipped.append({'date': d.isoformat(), 'seller': seller, 'buyer': buyer,
                                'truck': truck, 'contract': contract_str, 'reason': reason})

            parsed = parse_contract_number(contract_str)
            if not parsed:
                stats['skipped_nonstd'] += 1; skip('non_standard_contract'); continue
            seq, cyear = parsed

            ef = export_firms.get(seller)
            if ef is None:
                stats['skipped_no_firm'] += 1; skip('export_firm_not_found'); continue

            imp = imp_by_short.get(buyer.lower()) or imp_by_company.get(buyer.lower())
            if imp is None:
                stats['skipped_no_buyer'] += 1; skip('buyer_not_found'); continue

            # Match by plate, then the candidate whose date is closest within the
            # window (invoice date lags loading by ~1-2 weeks).
            WINDOW_DAYS = 21
            candidates = ship_by_plate.get(_norm_plate(truck), [])
            within = [(abs((sd - d).days), sh) for sd, sh in candidates
                      if abs((sd - d).days) <= WINDOW_DAYS]
            shipment = min(within, key=lambda x: x[0])[1] if within else None
            if shipment is None:
                stats['skipped_unmatched'] += 1
                skip('plate_not_in_db' if not candidates else 'date_outside_window')
                continue

            # Contract by canonical identity.
            contract = Contract.objects.filter(
                export_firm=ef, contract_year=cyear, seq=seq).first()
            if contract is None:
                stats['contracts_created'] += 1
            else:
                stats['contracts_linked'] += 1

            # Split presence.
            has_split = ShipmentFirmSplit.objects.filter(
                shipment=shipment, export_firm=ef).exists()
            if not has_split:
                stats['splits_created'] += 1

            sale_exists = ContractSale.objects.filter(
                shipment=shipment, export_firm=ef).exists()
            stats['sales_updated' if sale_exists else 'sales_created'] += 1

            actions.append(dict(shipment=shipment, ef=ef, imp=imp, seq=seq, cyear=cyear,
                                 number=contract_str, contract=contract, inv_no=inv_no,
                                 inv_date=d, qty=qty, usd=usd, passport=passport,
                                 has_split=has_split))

        self._report(stats, skipped, commit)

        if commit and actions:
            self._write(actions, season)
            self.stdout.write(self.style.SUCCESS(f'Committed {len(actions)} sales bridges.'))
        elif not commit:
            self.stdout.write(self.style.WARNING('DRY-RUN — no writes. Re-run with --commit.'))

    def _write(self, actions, season):
        with transaction.atomic():
            for a in actions:
                contract = a['contract']
                if contract is None:
                    contract = Contract.objects.create(
                        contract_number=a['number'], seq=a['seq'], contract_year=a['cyear'],
                        contract_type=Contract.TYPE_ONE_TIME, export_firm=a['ef'],
                        import_firm=a['imp'], season=season, start_date=a['inv_date'],
                        planned_trucks=1, status=Contract.STATUS_ACTIVE,
                        passport_sdelka=a['passport'] or '',
                    )
                if not a['has_split']:
                    ShipmentFirmSplit.objects.get_or_create(
                        shipment=a['shipment'], export_firm=a['ef'],
                        defaults=dict(weight_kg=a['qty'] or 0, amount_usd=a['usd']))
                ContractSale.objects.update_or_create(
                    shipment=a['shipment'], export_firm=a['ef'],
                    defaults=dict(contract=contract, import_firm=a['imp'],
                                  invoice_number=a['inv_no'], invoice_date=a['inv_date'],
                                  quantity_kg=a['qty'], total_usd=a['usd']))

    def _report(self, stats, skipped, commit):
        self.stdout.write('─' * 50)
        for k, v in stats.items():
            self.stdout.write(f'  {k:22s}: {v}')
        if skipped:
            out = Path(settings.BASE_DIR).parent / 'data' / 'import_2sales_june_skipped.csv'
            try:
                with open(out, 'w', newline='', encoding='utf-8') as fh:
                    w = csv.DictWriter(fh, fieldnames=list(skipped[0].keys()))
                    w.writeheader(); w.writerows(skipped)
                self.stdout.write(f'  skipped CSV: {out} ({len(skipped)} rows)')
            except OSError as e:
                self.stdout.write(f'  (could not write skipped CSV: {e})')
        self.stdout.write('─' * 50)
