"""Import master contracts from the 1-Contracts sheet of the season workbook.

Reads 43 rows from sheet '1-Contracts' in Export_contracts_2025-2026.xlsx and
creates Contract rows for those not already in the DB.  Idempotent: rows whose
contract_number already exists are silently skipped (no upsert in v1).

Usage:
    python manage.py import_contracts_master --dry-run
    python manage.py import_contracts_master --commit
    python manage.py import_contracts_master --commit --file /abs/path/to/file.xlsx

Flags:
    --dry-run   (default) Parse, validate, classify — NO writes.
    --commit    Real writes wrapped in transaction.atomic().
    --file PATH Override default Excel location.
"""
from __future__ import annotations

import datetime
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.contracts.models import Contract
from apps.core.models import ExportFirm, ImportFirm, Season


# Regex to extract DD.MM.YYYY from the trailing portion of a contract_number string.
_DATE_RE = re.compile(r'(\d{2}\.\d{2}\.\d{4})$')

BATCH_SIZE = 500


# ── helpers ────────────────────────────────────────────────────────────────────

def _clean(value) -> str | None:
    """Strip whitespace + trailing dot/comma; return None for empty cells."""
    if value is None:
        return None
    s = str(value).strip().rstrip('.,').strip()
    return s or None


def _safe_decimal(value) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _safe_int(value) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_start_date(contract_number: str) -> datetime.date | None:
    """Extract start_date from the trailing DD.MM.YYYY in the contract number."""
    m = _DATE_RE.search(contract_number)
    if not m:
        return None
    try:
        return datetime.datetime.strptime(m.group(1), '%d.%m.%Y').date()
    except ValueError:
        return None


# ── column constants (1-based) ─────────────────────────────────────────────────
# Header: [contract_number, seller_code, buyer, trucks, incoterm, qty_kg, sum_usd,
#          exported_qty, remainder_qty, exported_sum, remainder_sum,
#          payment, last_inv_no, sent_to_unk, ostatok, status]
COL_CONTRACT_NUMBER = 1
COL_SELLER_CODE = 2
COL_BUYER = 3
COL_TRUCKS = 4
COL_INCOTERM = 5
COL_PLANNED_QTY_KG = 6
COL_PLANNED_AMOUNT_USD = 7
COL_STATUS = 16  # col P: "OK" / other


class Command(BaseCommand):
    help = 'Import master contracts from 1-Contracts sheet (dry-run by default).'

    def add_arguments(self, parser):
        default_path = (
            Path(settings.BASE_DIR).parent / 'data' / 'Export_contracts_2025-2026.xlsx'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Validate and report counts — no writes (default mode).',
        )
        parser.add_argument(
            '--commit',
            action='store_true',
            default=False,
            help='Execute real writes (must be explicit; --dry-run takes precedence if both set).',
        )
        parser.add_argument(
            '--file',
            type=Path,
            default=default_path,
            metavar='PATH',
            help=f'Excel workbook path (default: {default_path})',
        )

    def handle(self, *args, **opts):
        path: Path = opts['file']
        # --dry-run beats --commit if both are supplied
        dry_run: bool = opts['dry_run'] or not opts['commit']

        if not path.exists():
            raise CommandError(f'File not found: {path}')

        self.stdout.write(f'Loading: {path}')
        wb = load_workbook(path, data_only=True, read_only=True)

        if '1-Contracts' not in wb.sheetnames:
            raise CommandError(
                f"Sheet '1-Contracts' not found. Available: {wb.sheetnames}"
            )

        ws = wb['1-Contracts']

        # ── Pre-load reference data ────────────────────────────────────────────
        export_firms: dict[str, ExportFirm] = {
            f.code: f for f in ExportFirm.objects.all()
        }
        import_firms_by_short: dict[str, ImportFirm] = {
            str(f.name_short or '').strip().lower(): f
            for f in ImportFirm.objects.all()
            if f.name_short
        }
        import_firms_by_company: dict[str, ImportFirm] = {
            str(f.name_company or '').strip().lower(): f
            for f in ImportFirm.objects.all()
            if f.name_company
        }
        season: Season | None = (
            Season.objects.filter(is_active=True).first()
            or Season.objects.order_by('-start_date').first()
        )

        if season is None:
            raise CommandError(
                'No Season row found in DB. '
                'Run seed_data or create a season before importing contracts.'
            )

        # ── Counters ──────────────────────────────────────────────────────────
        cnt_db_before = Contract.objects.count()
        cnt_rows = 0
        cnt_already_exist = 0
        cnt_skipped_fk = 0
        skipped_fk_rows: list[str] = []
        new_contracts: list[Contract] = []

        # ── Process rows ──────────────────────────────────────────────────────
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[COL_CONTRACT_NUMBER - 1] is None:
                continue

            raw_number = _clean(row[COL_CONTRACT_NUMBER - 1])
            if not raw_number:
                continue

            cnt_rows += 1

            # 1. Normalize contract number
            contract_number = raw_number

            # 2. Check if already in DB
            if Contract.objects.filter(contract_number__iexact=contract_number).exists():
                cnt_already_exist += 1
                continue

            # 3. Extract start_date from contract_number
            start_date = _parse_start_date(contract_number)

            # 4. Resolve ExportFirm
            seller_raw = _clean(row[COL_SELLER_CODE - 1]) or ''
            export_firm = export_firms.get(seller_raw)
            if export_firm is None:
                cnt_skipped_fk += 1
                skipped_fk_rows.append(
                    f'  row seller={seller_raw!r}  buyer=?  contract={contract_number!r}  '
                    f'reason=export_firm_not_found'
                )
                continue

            # 5. Resolve ImportFirm
            buyer_raw = _clean(row[COL_BUYER - 1]) or ''
            buyer_lower = buyer_raw.lower()
            import_firm = (
                import_firms_by_short.get(buyer_lower)
                or import_firms_by_company.get(buyer_lower)
            )
            if import_firm is None:
                cnt_skipped_fk += 1
                skipped_fk_rows.append(
                    f'  row seller={seller_raw!r}  buyer={buyer_raw!r}  '
                    f'contract={contract_number!r}  reason=import_firm_not_found'
                )
                continue

            # 6. Read planned fields
            planned_trucks = _safe_int(row[COL_TRUCKS - 1])
            incoterm = _clean(row[COL_INCOTERM - 1]) or ''
            planned_quantity_kg = _safe_decimal(row[COL_PLANNED_QTY_KG - 1])
            planned_amount_usd = _safe_decimal(row[COL_PLANNED_AMOUNT_USD - 1])
            status_raw = _clean(row[COL_STATUS - 1]) or ''
            status = Contract.STATUS_ACTIVE  # Excel sheet doesn't use DB status names

            # 7. Build Contract object (don't save yet)
            new_contracts.append(
                Contract(
                    contract_number=contract_number,
                    season=season,
                    export_firm=export_firm,
                    import_firm=import_firm,
                    contract_type='EXPORT',
                    incoterm=incoterm,
                    start_date=start_date,
                    planned_trucks=planned_trucks,
                    planned_quantity_kg=planned_quantity_kg,
                    planned_amount_usd=planned_amount_usd,
                    status=status,
                )
            )

        # ── Bulk insert ────────────────────────────────────────────────────────
        with transaction.atomic():
            if not dry_run and new_contracts:
                Contract.objects.bulk_create(new_contracts, batch_size=BATCH_SIZE)
            if dry_run:
                transaction.set_rollback(True)

        # ── Report ────────────────────────────────────────────────────────────
        self.stdout.write('')
        self.stdout.write(f'Contracts in DB before:       {cnt_db_before}')
        self.stdout.write(f'Rows in 1-Contracts sheet:    {cnt_rows}')
        self.stdout.write(f'Rows skipped (already exist): {cnt_already_exist}')
        self.stdout.write(
            f'Rows skipped (FK miss):       {cnt_skipped_fk}'
            + (' (with list below)' if skipped_fk_rows else '')
        )
        self.stdout.write(
            self.style.SUCCESS(
                f'Rows ready to import:         {len(new_contracts)}'
            )
        )

        if skipped_fk_rows:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('Skipped (FK miss):'))
            for msg in skipped_fk_rows:
                self.stdout.write(msg)

        if dry_run:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(
                '[DRY-RUN -- no writes performed. Re-run with --commit to write.]'
            ))
        else:
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS(
                f'Committed {len(new_contracts)} new Contract rows.'
            ))
