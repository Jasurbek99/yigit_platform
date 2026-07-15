"""Build the CMR overlay templates (``cmr_ru.xlsx`` / ``cmr_en.xlsx``).

Unlike the invoice/contract/letters — which are self-contained Word documents —
the office CMR is a **print overlay**: data positioned to print on top of the
pre-printed official CMR (red 24-box) form. The source ``CMR RU`` / ``CMR EN``
sheets carry that exact geometry (column widths, row heights, A4 @ 60% scale,
merges) tuned to register on the physical form, with the cell values driven by
VLOOKUPs into other sheets and a block of green/yellow helper input cells off to
the right (columns O+).

Reproducing that grid in python-docx would be lossy by construction, so we keep
the Excel geometry and fill cells by coordinate at render time (see
``document_render.render_xlsx`` + ``document_context.build_cmr_overlay``).

This builder strips the source sheet down to a clean template:
  * every data / formula cell is blanked (no dependency on the other sheets),
  * the helper input columns (O onward) are cleared,
  * only the fixed unit labels (``Брутто:`` / ``кг.`` / ``вес поддона`` …) and the
    grid geometry survive.

Run once to (re)create the committed templates:

    python -m apps.contracts.document_templates.build_cmr_xlsx [SOURCE_XLSX]
"""
from pathlib import Path
import sys

import openpyxl

OUT_DIR = Path(__file__).resolve().parent
REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_SOURCE = REPO_ROOT / 'data' / 'Export_contracts_2025-2026.xlsx'

# First printed CMR column is A; everything from column O (index 15) rightward is
# the helper input block and must never survive into the template.
FIRST_HELPER_COL = 15
MAX_ROW = 60

# Cells to KEEP as fixed labels per source sheet (coord → its constant text is
# preserved). Everything else with a value in columns A–N is blanked.
KEEP_LABELS = {
    'CMR RU': {
        'L26', 'N26', 'I27', 'H28', 'H29', 'M27', 'M28', 'M29', 'D46',
    },
    'CMR EN': {
        'L26', 'N26', 'I27', 'G28', 'G29', 'M27', 'M28', 'M29', 'D46',
    },
}

OUT_NAME = {'CMR RU': 'cmr_ru.xlsx', 'CMR EN': 'cmr_en.xlsx'}


def _clean_sheet(ws, keep: set[str]) -> None:
    """Blank all data/formula cells and the helper columns, keeping fixed labels."""
    for row in ws.iter_rows(min_row=1, max_row=MAX_ROW):
        for cell in row:
            if cell.value in (None, ''):
                continue
            if cell.column >= FIRST_HELPER_COL:
                cell.value = None
            elif cell.coordinate not in keep:
                cell.value = None


def build(source: Path, sheet_name: str) -> Path:
    # data_only so kept label cells resolve to plain strings, not formulas.
    wb = openpyxl.load_workbook(source, data_only=True)
    for name in list(wb.sheetnames):
        if name != sheet_name:
            del wb[name]
    ws = wb[sheet_name]
    _clean_sheet(ws, KEEP_LABELS[sheet_name])
    # Constrain the print to the CMR content grid (cols A–N) so the 60%-scale
    # overlay lands on one page instead of trailing empty rows/cols. Row 54 is
    # included because the EN sheet's vehicle-plate cell (F54) sits there — the RU
    # sheet ends at row 53, so its row 54 is simply blank.
    ws.print_area = 'A1:N54'
    out = OUT_DIR / OUT_NAME[sheet_name]
    wb.save(out)
    return out


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        raise SystemExit(f'Source workbook not found: {source}')
    for sheet_name in ('CMR RU', 'CMR EN'):
        print(f'wrote {build(source, sheet_name)}')


if __name__ == '__main__':
    main()
