"""Upsert ExportFirm and ImportFirm rows from data/export_import_firms.xlsx.

The DB currently has many incorrect/placeholder firm rows (auto-created from
shipment imports). This command treats the Excel workbook as the source of
truth: existing rows matched by natural key are UPDATED, missing rows are
INSERTED. Rows already in DB that are NOT in the Excel are left alone.

Natural keys:
  - ExportFirm: `code` (rows with no code in Excel are skipped)
  - ImportFirm: stripped `code` if present, else `name_company`

Usage:
    python manage.py import_firms_from_excel
    python manage.py import_firms_from_excel --dry-run
    python manage.py import_firms_from_excel --file path/to/file.xlsx
"""
from __future__ import annotations

from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.core.models import ExportFirm, ImportFirm


# 1-based Excel column positions (see analysis in CHANGELOG).
EXPORT_COLS = {
    'code': 2,
    'name_ru': 3,
    'address_ru': 4,
    'bank_details_ru': 5,
    'director': 6,
    'name_en': 8,
    'address_en': 9,
    'bank_details_en': 10,
    'name_tk': 12,
    'address_tk': 13,
    'bank_details_tk': 14,
}

IMPORT_COLS = {
    'code': 2,
    'name_company': 3,
    'address': 4,
    'bank_details': 5,
    'name_short': 6,
}

# Per-field max lengths (must match firms.py); used to truncate over-long cells
# rather than raising. Truncations are reported.
EXPORT_MAX = {
    'code': 20,
    'name_tk': 200, 'name_ru': 200, 'name_en': 200,
    'address_tk': 500, 'address_ru': 500, 'address_en': 500,
    'bank_details_tk': 1000, 'bank_details_ru': 1000, 'bank_details_en': 1000,
    'director': 200,
}

IMPORT_MAX = {
    'code': 50,
    'name_company': 300,
    'name_short': 100,
    'address': 500,
    'bank_details': 1000,
}


def _clean(value) -> str | None:
    """Normalize a cell: cast to str, strip whitespace + trailing comma/underscores."""
    if value is None:
        return None
    s = str(value).strip()
    # Excel director cells use trailing underscores as a signature line filler
    s = s.rstrip('_').rstrip()
    # Some name_tk values end with stray comma
    s = s.rstrip(',').rstrip()
    return s or None


def _truncate(value: str | None, max_len: int, *, field: str, row_label: str,
              warnings: list[str]) -> str | None:
    if value is None:
        return None
    if len(value) > max_len:
        warnings.append(
            f'  truncated {row_label} {field}: {len(value)} -> {max_len} chars'
        )
        return value[:max_len]
    return value


class Command(BaseCommand):
    help = 'Upsert ExportFirm + ImportFirm rows from export_import_firms.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('--dry-run', action='store_true',
                            help='Show what would change, write nothing.')
        default_path = Path(settings.BASE_DIR).parent / 'data' / 'export_import_firms.xlsx'
        parser.add_argument('--file', type=Path, default=default_path,
                            help=f'Excel path (default: {default_path})')

    def handle(self, *args, **opts):
        path: Path = opts['file']
        dry_run: bool = opts['dry_run']

        if not path.exists():
            raise CommandError(f'File not found: {path}')

        self.stdout.write(f'Loading: {path}')
        # read_only=False so ws.max_row is accurate (files are small)
        wb = load_workbook(path, data_only=True)

        if 'export' not in wb.sheetnames or 'import' not in wb.sheetnames:
            raise CommandError(
                f"Expected sheets 'export' and 'import'; got: {wb.sheetnames}")

        warnings: list[str] = []

        with transaction.atomic():
            export_stats = self._upsert_export(wb['export'], dry_run, warnings)
            import_stats = self._upsert_import(wb['import'], dry_run, warnings)
            if dry_run:
                self.stdout.write(self.style.WARNING(
                    'DRY RUN — rolling back transaction.'))
                transaction.set_rollback(True)

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=== ExportFirm ==='))
        self.stdout.write(f'  created : {export_stats["created"]}')
        self.stdout.write(f'  updated : {export_stats["updated"]}')
        self.stdout.write(f'  skipped : {export_stats["skipped"]} (no code)')
        self.stdout.write(self.style.SUCCESS('=== ImportFirm ==='))
        self.stdout.write(f'  created : {import_stats["created"]}')
        self.stdout.write(f'  updated : {import_stats["updated"]}')
        self.stdout.write(f'  skipped : {import_stats["skipped"]} (no key)')

        if warnings:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(
                f'{len(warnings)} warning(s):'))
            for w in warnings:
                self.stdout.write(w)

    # ── ExportFirm ────────────────────────────────────────────────────────────
    def _upsert_export(self, ws, dry_run: bool, warnings: list[str]) -> dict:
        stats = {'created': 0, 'updated': 0, 'skipped': 0}

        for r in range(2, ws.max_row + 1):
            row = {field: _clean(ws.cell(row=r, column=col).value)
                   for field, col in EXPORT_COLS.items()}

            # Fully empty trailing row — stop iterating
            if not any(row.values()):
                continue

            code = row['code']
            if not code:
                # Row has data but no natural key — skip with warning.
                stats['skipped'] += 1
                warnings.append(
                    f'  export row {r}: no code; skipped '
                    f'(name_ru={row["name_ru"]!r})')
                continue
            code = code[:EXPORT_MAX['code']]

            # name_tk is required (blank=False, null=False) — fall back to name_ru
            # for rows with code but missing tk name.
            name_tk = row['name_tk'] or row['name_ru'] or row['name_en']
            if not name_tk:
                warnings.append(
                    f'  export row {r} code={code!r}: no name in any language — skipped')
                stats['skipped'] += 1
                continue

            row_label = f'export row {r} [{code}]'
            payload = {
                'name_tk': _truncate(name_tk, EXPORT_MAX['name_tk'],
                                     field='name_tk', row_label=row_label,
                                     warnings=warnings),
                'name_ru': _truncate(row['name_ru'], EXPORT_MAX['name_ru'],
                                     field='name_ru', row_label=row_label,
                                     warnings=warnings),
                'name_en': _truncate(row['name_en'], EXPORT_MAX['name_en'],
                                     field='name_en', row_label=row_label,
                                     warnings=warnings),
                'address_tk': _truncate(row['address_tk'], EXPORT_MAX['address_tk'],
                                        field='address_tk', row_label=row_label,
                                        warnings=warnings),
                'address_ru': _truncate(row['address_ru'], EXPORT_MAX['address_ru'],
                                        field='address_ru', row_label=row_label,
                                        warnings=warnings),
                'address_en': _truncate(row['address_en'], EXPORT_MAX['address_en'],
                                        field='address_en', row_label=row_label,
                                        warnings=warnings),
                'bank_details_tk': _truncate(row['bank_details_tk'], EXPORT_MAX['bank_details_tk'],
                                             field='bank_details_tk', row_label=row_label,
                                             warnings=warnings),
                'bank_details_ru': _truncate(row['bank_details_ru'], EXPORT_MAX['bank_details_ru'],
                                             field='bank_details_ru', row_label=row_label,
                                             warnings=warnings),
                'bank_details_en': _truncate(row['bank_details_en'], EXPORT_MAX['bank_details_en'],
                                             field='bank_details_en', row_label=row_label,
                                             warnings=warnings),
                'director': _truncate(row['director'], EXPORT_MAX['director'],
                                      field='director', row_label=row_label,
                                      warnings=warnings),
            }

            existing = ExportFirm.objects.filter(code=code).first()
            if existing:
                changed = self._diff_and_apply(existing, payload)
                if changed:
                    if not dry_run:
                        existing.save(update_fields=list(payload.keys()))
                    stats['updated'] += 1
                    self.stdout.write(
                        f'  UPDATE  ExportFirm[{code}] '
                        f'({len(changed)} field(s): {", ".join(changed)})')
            else:
                if not dry_run:
                    ExportFirm.objects.create(code=code, **payload)
                stats['created'] += 1
                self.stdout.write(f'  CREATE  ExportFirm[{code}] {payload["name_tk"]}')

        return stats

    # ── ImportFirm ────────────────────────────────────────────────────────────
    def _upsert_import(self, ws, dry_run: bool, warnings: list[str]) -> dict:
        stats = {'created': 0, 'updated': 0, 'skipped': 0}

        for r in range(2, ws.max_row + 1):
            row = {field: _clean(ws.cell(row=r, column=col).value)
                   for field, col in IMPORT_COLS.items()}

            # Fully empty trailing row — skip silently
            if not any(row.values()):
                continue

            code = row['code']
            name_company = row['name_company']

            if not code and not name_company:
                # Row has data only in unmapped columns
                stats['skipped'] += 1
                continue

            if not name_company:
                warnings.append(
                    f'  import row {r} code={code!r}: no name_company — skipped')
                stats['skipped'] += 1
                continue

            if code:
                code = code[:IMPORT_MAX['code']]

            row_label = f'import row {r} [{code or name_company[:30]}]'
            payload = {
                'name_company': _truncate(name_company, IMPORT_MAX['name_company'],
                                          field='name_company', row_label=row_label,
                                          warnings=warnings),
                'name_short': _truncate(row['name_short'], IMPORT_MAX['name_short'],
                                        field='name_short', row_label=row_label,
                                        warnings=warnings),
                'address': _truncate(row['address'], IMPORT_MAX['address'],
                                     field='address', row_label=row_label,
                                     warnings=warnings),
                'bank_details': _truncate(row['bank_details'], IMPORT_MAX['bank_details'],
                                          field='bank_details', row_label=row_label,
                                          warnings=warnings),
            }

            existing = None
            if code:
                existing = ImportFirm.objects.filter(code=code).first()
            if existing is None:
                # Fallback: match by exact name_company (covers code=NULL rows
                # and pre-existing rows imported from shipment data with no code)
                existing = ImportFirm.objects.filter(
                    name_company=payload['name_company']).first()

            if existing:
                update_payload = dict(payload)
                # Only overwrite code if it would set a real value (don't blank
                # a populated code with NULL from Excel)
                if code and existing.code != code:
                    update_payload['code'] = code
                changed = self._diff_and_apply(existing, update_payload)
                if changed:
                    if not dry_run:
                        existing.save(update_fields=list(update_payload.keys()))
                    stats['updated'] += 1
                    self.stdout.write(
                        f'  UPDATE  ImportFirm[{existing.code or existing.name_company[:20]}] '
                        f'({len(changed)} field(s): {", ".join(changed)})')
            else:
                create_payload = dict(payload)
                if code:
                    create_payload['code'] = code
                if not dry_run:
                    ImportFirm.objects.create(**create_payload)
                stats['created'] += 1
                self.stdout.write(
                    f'  CREATE  ImportFirm[{code or "-"}] {payload["name_company"][:50]}')

        return stats

    # ── helpers ───────────────────────────────────────────────────────────────
    @staticmethod
    def _diff_and_apply(instance, payload: dict) -> list[str]:
        """Set attrs from payload only where they differ; return changed field names."""
        changed = []
        for field, new_value in payload.items():
            current = getattr(instance, field)
            if current != new_value:
                setattr(instance, field, new_value)
                changed.append(field)
        return changed
