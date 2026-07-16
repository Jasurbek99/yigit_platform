"""Parse a weightmaster loading-detail Excel report into pallet manifest rows.

The weightmaster (Artykow Maksat at Kaka) produces one .xlsx per truck: a
per-pallet weighing table. This module reads that table and resolves each row's
crate type, variety and sub-block against reference data, so the pallet manifest
can be pre-filled instead of hand-keyed. See project doc: weightmaster feature.

Template (verified from weightmaster_report_10AP116.xlsx, first sheet, row 1 =
header, data rows 2..N until column A is blank):

    A  PALET №                  "Palet 1"        -> pallet_number
    B  DOLY AGRAM               474              -> gross_weight_kg
    C  1 GAP AGRAM              0.543            -> crate unit weight (resolves crate_type)
    D  GAP SANY                 64               -> crate_count
    F  POLET AGRAM              7.5              -> pallet_weight_kg
    G  GOŞUNDYLAR               4                -> additions_kg
    I  POMIDORYŇ GÖRNÜŞI        "MIDELICE"       -> variety (by name)
    J  KODLAMA                  "10AP116-F01"    -> load code prefix (10AP116)
    K  BÖLÜMI                   "F2"             -> sub_block (by code)
    M  ÝYGYLAN SENESI           "10,04,2026"     -> harvest date (DD,MM,YYYY text)

Nothing is silently dropped: a value that does not resolve to reference data is
returned with a null id plus a warning carrying the raw text, so a human fixes it
in the grid before saving.
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl

logger = logging.getLogger(__name__)

# Column letters -> 1-based indices for the fixed template.
_COL = {
    'pallet': 1,   # A
    'gross': 2,    # B
    'crate_w': 3,  # C
    'count': 4,    # D
    'pallet_w': 6, # F
    'additions': 7,  # G
    'variety': 9,  # I
    'kodlama': 10, # J
    'block': 11,   # K
    'harvest': 13, # M
}

_PALLET_RE = re.compile(r'(\d+)')
_HEADER_HINT = 'PALET'  # A1 of a valid template starts with this


@dataclass
class ParsedPalletRow:
    """One parsed pallet row, resolved against reference data where possible."""

    pallet_number: int
    gross_weight_kg: Decimal
    crate_count: int
    pallet_weight_kg: Decimal
    additions_kg: Decimal
    # Resolved FK ids (None when the raw value did not match reference data)
    crate_type: int | None
    crate_type_name: str
    variety: int | None
    variety_name: str
    sub_block: int | None
    sub_block_code: str


@dataclass
class ParsedWarning:
    """A row-level resolution problem the user must fix before saving."""

    row: int          # 1-based Excel row number
    pallet_number: int | None
    field: str        # 'crate_type' | 'variety' | 'sub_block' | 'row'
    raw: str          # verbatim cell text that failed to resolve
    message: str


@dataclass
class ParsedWeightmaster:
    """Full parse result: rows to preview, warnings, and a summary header."""

    rows: list[ParsedPalletRow] = field(default_factory=list)
    warnings: list[ParsedWarning] = field(default_factory=list)
    load_code: str = ''
    harvest_date: str | None = None  # ISO date string, or None if unparseable
    total_gross_kg: Decimal = Decimal('0')
    total_net_kg: Decimal = Decimal('0')


class WeightmasterParseError(ValueError):
    """Raised when the file is not a readable weightmaster template."""


def _to_decimal(value, default: Decimal = Decimal('0')) -> Decimal | None:
    """Coerce an Excel cell value to Decimal, or None if blank/invalid."""
    if value is None or value == '':
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _parse_harvest_date(raw) -> str | None:
    """Parse the 'DD,MM,YYYY' text date into an ISO string, or None."""
    if not raw:
        return None
    parts = re.split(r'[,./-]', str(raw).strip())
    if len(parts) != 3:
        return None
    try:
        day, month, year = (int(p) for p in parts)
        return date(year, month, day).isoformat()
    except (ValueError, TypeError):
        return None


def parse_weightmaster_workbook(file_obj) -> ParsedWeightmaster:
    """Parse a weightmaster .xlsx into pallet rows resolved against reference data.

    Args:
        file_obj: An uploaded file (Django UploadedFile) or path/bytes openpyxl
            can load.

    Returns:
        ParsedWeightmaster with rows, warnings and a summary header.

    Raises:
        WeightmasterParseError: If the workbook cannot be opened or the first
            sheet does not look like the weightmaster template.
    """
    try:
        # Not read_only: in read-only mode ws.max_row is read from the <dimension>
        # tag and can under-report, silently dropping trailing pallets. These
        # files are tiny (~100 rows), so full load is cheap and reliable.
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of errors on bad files
        raise WeightmasterParseError(f'Could not read the Excel file: {exc}') from exc

    ws = wb.worksheets[0]
    header = ws.cell(row=1, column=_COL['pallet']).value
    if not header or _HEADER_HINT not in str(header).upper():
        raise WeightmasterParseError(
            'This does not look like a weightmaster report '
            '(cell A1 should be the "PALET №" header).'
        )

    crate_by_weight, varieties_by_name, blocks_by_code = _load_reference_maps()

    result = ParsedWeightmaster()
    for excel_row in range(2, ws.max_row + 1):
        pallet_cell = ws.cell(row=excel_row, column=_COL['pallet']).value
        if pallet_cell is None or str(pallet_cell).strip() == '':
            break  # blank column A marks the end of the pallet table

        match = _PALLET_RE.search(str(pallet_cell))
        if not match:
            break  # a non-pallet row (e.g. totals) — stop reading

        _parse_one_row(
            ws, excel_row, int(match.group(1)),
            crate_by_weight, varieties_by_name, blocks_by_code, result,
        )

    result.load_code = _extract_load_code(ws)
    wb.close()
    logger.info(
        'Weightmaster parsed: %d pallets, %d warnings, load_code=%s',
        len(result.rows), len(result.warnings), result.load_code,
    )
    return result


def _parse_one_row(
    ws, excel_row: int, pallet_number: int,
    crate_by_weight: dict, varieties_by_name: dict, blocks_by_code: dict,
    result: ParsedWeightmaster,
) -> None:
    """Parse and resolve a single pallet row, appending row + any warnings."""
    def cell(key):
        return ws.cell(row=excel_row, column=_COL[key]).value

    # default=None so a BLANK gross cell trips the skip-guard below, rather than
    # defaulting to 0 and producing a nonsensical negative net with no warning.
    gross = _to_decimal(cell('gross'), default=None)
    pallet_w = _to_decimal(cell('pallet_w'))
    additions = _to_decimal(cell('additions'))
    crate_w = _to_decimal(cell('crate_w'), default=None)
    count_raw = cell('count')

    try:
        count = int(count_raw) if count_raw not in (None, '') else 0
    except (TypeError, ValueError):
        count = 0

    if gross is None or count <= 0:
        result.warnings.append(ParsedWarning(
            row=excel_row, pallet_number=pallet_number, field='row',
            raw=f'gross={cell("gross")!r} count={count_raw!r}',
            message='Row skipped: missing or invalid gross weight / crate count.',
        ))
        return

    variety_raw = (str(cell('variety')).strip() if cell('variety') else '')
    block_raw = (str(cell('block')).strip() if cell('block') else '')
    pallet_w_val = pallet_w if pallet_w is not None else Decimal('0')
    additions_val = additions if additions is not None else Decimal('0')
    crate_total = (crate_w * count) if crate_w is not None else Decimal('0')
    net = gross - crate_total - pallet_w_val - additions_val
    result.total_gross_kg += gross
    result.total_net_kg += net
    if result.harvest_date is None:
        result.harvest_date = _parse_harvest_date(cell('harvest'))

    crate = crate_by_weight.get(crate_w.quantize(Decimal('0.001'))) if crate_w is not None else None
    if crate is None:
        result.warnings.append(ParsedWarning(
            row=excel_row, pallet_number=pallet_number, field='crate_type',
            raw=str(cell('crate_w')),
            message=f'No crate type with weight {cell("crate_w")} kg — pick one manually.',
        ))

    variety = varieties_by_name.get(variety_raw.lower())
    if variety is None and variety_raw:
        result.warnings.append(ParsedWarning(
            row=excel_row, pallet_number=pallet_number, field='variety',
            raw=variety_raw,
            message=f'Variety "{variety_raw}" not found — pick one manually.',
        ))

    block = blocks_by_code.get(block_raw.lower())
    if block is None and block_raw:
        result.warnings.append(ParsedWarning(
            row=excel_row, pallet_number=pallet_number, field='sub_block',
            raw=block_raw,
            message=f'Block "{block_raw}" not found — pick one manually.',
        ))

    result.rows.append(ParsedPalletRow(
        pallet_number=pallet_number,
        gross_weight_kg=gross,
        crate_count=count,
        pallet_weight_kg=pallet_w_val,
        additions_kg=additions_val,
        crate_type=crate.id if crate else None,
        crate_type_name=crate.name if crate else '',
        variety=variety.id if variety else None,
        variety_name=variety.name if variety else variety_raw,
        sub_block=block.id if block else None,
        sub_block_code=block.code if block else block_raw,
    ))


def _load_reference_maps() -> tuple[dict, dict, dict]:
    """Load reference data into lookup dicts (single query each, no N+1)."""
    from apps.core.models import CrateType, GreenhouseBlock, TomatoVariety

    crate_by_weight = {
        c.weight_kg.quantize(Decimal('0.001')): c
        for c in CrateType.objects.all()
    }
    varieties_by_name = {v.name.lower(): v for v in TomatoVariety.objects.all()}
    blocks_by_code = {b.code.lower(): b for b in GreenhouseBlock.objects.all()}
    return crate_by_weight, varieties_by_name, blocks_by_code


def _extract_load_code(ws) -> str:
    """Pull the load/waybill code (e.g. '10AP116') from the first KODLAMA cell."""
    raw = ws.cell(row=2, column=_COL['kodlama']).value
    if not raw:
        return ''
    # KODLAMA is like "10AP116-F01"; the load code is the part before "-".
    return str(raw).split('-')[0].strip()
