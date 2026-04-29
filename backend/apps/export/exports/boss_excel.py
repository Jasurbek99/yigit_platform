"""Excel writers for Boss Dashboard report sections.

Each section calls the matching aggregator from services/boss_analytics.py and
writes a single .xlsx workbook. Returns bytes ready for HttpResponse.

The aggregators are the single source of truth for the numbers — Excel and JSON
share one code path, so there is no risk of divergence between what's on screen
and what's downloaded.
"""
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.export.services.boss_analytics import (
    _aggregate_summary,
    _aggregate_revenue,
    _aggregate_route_pnl,
    _aggregate_quota_grid,
    _aggregate_blocks_heatmap,
    _aggregate_top_customers,
    _aggregate_compliance,
    _aggregate_risk_matrix,
    _aggregate_production,
    _aggregate_export_market,
)


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_HEADER_FILL = PatternFill('solid', fgColor='1677FF')
_TITLE_FONT = Font(bold=True, size=14)
_TOTAL_FONT = Font(bold=True)
_TOTAL_FILL = PatternFill('solid', fgColor='F0F5FF')
_CENTER = Alignment(horizontal='center', vertical='center')


def _write_title(ws, title: str, period_label: str) -> int:
    """Write a section title and period label. Return next row index."""
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=period_label).font = Font(italic=True, color='8C8C8C')
    return 4


def _write_header(ws, row: int, headers: list[str]) -> int:
    """Write a styled header row. Return next row index."""
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    return row + 1


def _autosize(ws, max_width: int = 40) -> None:
    """Approximate column auto-sizing based on cell content length."""
    for col in ws.columns:
        column_letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[column_letter].width = min(max(longest + 2, 12), max_width)


def _period_label(from_date: date, to_date: date) -> str:
    return f'{from_date.isoformat()} → {to_date.isoformat()}'


# ---------------------------------------------------------------------------
# Section writers
# ---------------------------------------------------------------------------

def _write_monthly(wb: Workbook, from_date: date, to_date: date) -> None:
    """KPI summary + revenue trend for the period."""
    summary = _aggregate_summary(from_date, to_date)
    revenue = _aggregate_revenue(from_date, to_date)

    ws = wb.active
    ws.title = 'KPIs'
    row = _write_title(ws, 'Aýlyk hasabat — KPI', _period_label(from_date, to_date))

    row = _write_header(ws, row, ['KPI', 'Bahasy', 'Goşmaça'])
    kpis = summary  # dict of dicts
    label_map = {
        'revenue':      'Möwsüm girdejisi (USD)',
        'margin':       'Margin (USD)',
        'debt':         'Bergi (USD) — DEMO',
        'today_loaded': 'Bu gün ýüklendi',
        'in_transit':   'Ýolda maşyn',
        'quota_used':   'Kwota ulanyldy (%)',
    }
    for key, label in label_map.items():
        kpi = kpis.get(key, {})
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=float(kpi.get('value') or 0))
        extra = kpi.get('delta_pct') or kpi.get('plan') or kpi.get('total_season') or kpi.get('firms_total') or ''
        ws.cell(row=row, column=3, value=str(extra))
        row += 1
    _autosize(ws)

    # Revenue trend on a second sheet
    ws2 = wb.create_sheet('Revenue trend')
    row = _write_title(ws2, 'Girdeji — hepde-hepde', _period_label(from_date, to_date))
    row = _write_header(ws2, row, ['Hepde başlangyjy', 'Häzirki möwsüm (USD)', 'Geçen möwsüm (USD)'])
    cur = {r['week_start']: float(r['total_usd']) for r in revenue.get('current_season', [])}
    prev = {r['week_start']: float(r['total_usd']) for r in revenue.get('previous_season', [])}
    for week in sorted(set(cur) | set(prev)):
        ws2.cell(row=row, column=1, value=week)
        ws2.cell(row=row, column=2, value=cur.get(week, 0))
        ws2.cell(row=row, column=3, value=prev.get(week, 0))
        row += 1
    _autosize(ws2)


def _write_firms(wb: Workbook, from_date: date, to_date: date) -> None:
    """All firms with quota usage + risk level."""
    quota_rows = _aggregate_quota_grid()
    risk_rows = {r['firm_id']: r for r in _aggregate_risk_matrix()}

    ws = wb.active
    ws.title = 'Firmalar'
    row = _write_title(ws, 'Firma boýunça hasabat', _period_label(from_date, to_date))
    row = _write_header(ws, row, ['Firma', 'Kwota %', 'Derejesi', 'Bergi (DEMO)', 'Bank kredit (DEMO)', 'Risk'])
    for q in quota_rows:
        risk = risk_rows.get(q.get('firm_id'), {})
        ws.cell(row=row, column=1, value=q.get('firm_name', ''))
        ws.cell(row=row, column=2, value=float(q.get('used_pct') or 0))
        ws.cell(row=row, column=3, value=q.get('level', ''))
        debt = risk.get('debt_usd') or {}
        credit = risk.get('bank_credit_usd') or {}
        ws.cell(row=row, column=4, value=float(debt.get('value') or 0))
        ws.cell(row=row, column=5, value=float(credit.get('value') or 0))
        ws.cell(row=row, column=6, value=risk.get('risk_level', ''))
        row += 1
    _autosize(ws)


def _write_routes(wb: Workbook, from_date: date, to_date: date) -> None:
    """Route P&L breakdown."""
    rows = _aggregate_route_pnl(from_date, to_date)

    ws = wb.active
    ws.title = 'Marşrutlar'
    row = _write_title(ws, 'Marşrut girdejililigi (P&L)', _period_label(from_date, to_date))
    row = _write_header(ws, row, ['Ýurt', 'Şäher', 'Maşyn', 'Girdeji (USD)', 'Çykdajy (USD)', 'Marja (USD)', 'Marja %'])
    for r in rows:
        ws.cell(row=row, column=1, value=r.get('country_name', ''))
        ws.cell(row=row, column=2, value=r.get('city', ''))
        ws.cell(row=row, column=3, value=r.get('trucks', 0))
        ws.cell(row=row, column=4, value=float(r.get('revenue_usd') or 0))
        ws.cell(row=row, column=5, value=float(r.get('cost_usd') or 0))
        ws.cell(row=row, column=6, value=float(r.get('margin_usd') or 0))
        ws.cell(row=row, column=7, value=float(r.get('margin_pct') or 0))
        row += 1
    _autosize(ws)


def _write_blocks(wb: Workbook, from_date: date, to_date: date) -> None:
    """Greenhouse plan vs actual + production results + export-market split.

    Içerki Bazar and Sowgatlyk columns are intentionally absent (v1 scope).
    """
    heatmap = _aggregate_blocks_heatmap(from_date, to_date)
    production_daily = _aggregate_production('daily', from_date, to_date)
    production_seasonal = _aggregate_production('seasonal', from_date, to_date)
    export_market = _aggregate_export_market(from_date, to_date)

    # Sheet 1: heatmap
    ws = wb.active
    ws.title = 'Heatmap (7 gün)'
    row = _write_title(ws, 'Ýyladyşhanalar — hasyl ýagdaýy', _period_label(from_date, to_date))
    row = _write_header(ws, row, ['Blok', 'Plan KG', 'Fakt KG', '% planyň', 'Ýagdaý'])
    for r in heatmap:
        ws.cell(row=row, column=1, value=r.get('block_code', ''))
        ws.cell(row=row, column=2, value=float(r.get('plan_kg') or 0))
        ws.cell(row=row, column=3, value=float(r.get('actual_kg') or 0))
        ws.cell(row=row, column=4, value=float(r.get('pct') or 0))
        ws.cell(row=row, column=5, value=r.get('color_band', ''))
        row += 1
    _autosize(ws)

    # Sheet 2: daily production results
    ws2 = wb.create_sheet('Günlük önümçilik')
    row = _write_title(ws2, 'Günlük önümçilik netijeleri', _period_label(from_date, to_date))
    row = _write_header(ws2, row, ['Ýyladyşhana', 'Meýilleşdirilen-KG', 'Yerine Yetirilen-KG', '% (gün)', 'Aýlyk plan-KG', 'Aýlyk fakt-KG', '% (aý)'])
    for r in production_daily:
        ws2.cell(row=row, column=1, value=r.get('block_code', ''))
        ws2.cell(row=row, column=2, value=float(r.get('plan_kg') or 0))
        ws2.cell(row=row, column=3, value=float(r.get('actual_kg') or 0))
        ws2.cell(row=row, column=4, value=float(r.get('pct') or 0))
        ws2.cell(row=row, column=5, value=float(r.get('monthly_plan_kg') or 0))
        ws2.cell(row=row, column=6, value=float(r.get('monthly_actual_kg') or 0))
        ws2.cell(row=row, column=7, value=float(r.get('monthly_pct') or 0))
        row += 1
    _autosize(ws2)

    # Sheet 3: seasonal production results
    ws3 = wb.create_sheet('Möwsümleýin önümçilik')
    row = _write_title(ws3, 'Möwsümleýin önümçilik netijeleri', _period_label(from_date, to_date))
    row = _write_header(ws3, row, ['Ýyladyşhana', 'Meýilleşdirilen-KG', 'Yerine Yetirilen-KG', '% (möwsüm)', 'Aýlyk plan-KG', 'Aýlyk fakt-KG', '% (aý)'])
    for r in production_seasonal:
        ws3.cell(row=row, column=1, value=r.get('block_code', ''))
        ws3.cell(row=row, column=2, value=float(r.get('plan_kg') or 0))
        ws3.cell(row=row, column=3, value=float(r.get('actual_kg') or 0))
        ws3.cell(row=row, column=4, value=float(r.get('pct') or 0))
        ws3.cell(row=row, column=5, value=float(r.get('monthly_plan_kg') or 0))
        ws3.cell(row=row, column=6, value=float(r.get('monthly_actual_kg') or 0))
        ws3.cell(row=row, column=7, value=float(r.get('monthly_pct') or 0))
        row += 1
    _autosize(ws3)

    # Sheet 4: export-market by block (Daşarky Bazar only — Içerki/Sowgatlyk excluded by design)
    ws4 = wb.create_sheet('Daşarky Bazar')
    row = _write_title(ws4, 'Daşarky Bazar — blok boýunça', _period_label(from_date, to_date))
    row = _write_header(ws4, row, ['Ýyladyşhana', 'Daşarky Bazar (KG)', 'Daşarky Bazar (%)'])
    for r in export_market:
        ws4.cell(row=row, column=1, value=r.get('block_code', ''))
        ws4.cell(row=row, column=2, value=float(r.get('export_kg') or 0))
        ws4.cell(row=row, column=3, value=float(r.get('export_pct') or 0))
        row += 1
    _autosize(ws4)


def _write_seasons(wb: Workbook, from_date: date, to_date: date) -> None:
    """Current season vs previous season revenue comparison."""
    revenue = _aggregate_revenue(from_date, to_date)
    summary = _aggregate_summary(from_date, to_date)

    ws = wb.active
    ws.title = 'Möwsüm deňeşdirme'
    row = _write_title(ws, 'Möwsüm deňeşdirme', _period_label(from_date, to_date))

    rev_kpi = summary.get('revenue', {})
    ws.cell(row=row, column=1, value='Häzirki möwsüm girdejisi (USD)').font = _TOTAL_FONT
    ws.cell(row=row, column=2, value=float(rev_kpi.get('value') or 0))
    row += 1
    ws.cell(row=row, column=1, value='Üýtgeme (%) geçen ýyl bilen').font = _TOTAL_FONT
    ws.cell(row=row, column=2, value=float(rev_kpi.get('delta_pct') or 0))
    row += 2

    row = _write_header(ws, row, ['Hepde başlangyjy', 'Häzirki (USD)', 'Geçen (USD)', 'Tapawut (USD)'])
    cur = {r['week_start']: float(r['total_usd']) for r in revenue.get('current_season', [])}
    prev = {r['week_start']: float(r['total_usd']) for r in revenue.get('previous_season', [])}
    for week in sorted(set(cur) | set(prev)):
        c = cur.get(week, 0)
        p = prev.get(week, 0)
        ws.cell(row=row, column=1, value=week)
        ws.cell(row=row, column=2, value=c)
        ws.cell(row=row, column=3, value=p)
        ws.cell(row=row, column=4, value=c - p)
        row += 1
    _autosize(ws)


def _write_audit(wb: Workbook, from_date: date, to_date: date) -> None:
    """Recent shipment status transitions — who, what, when."""
    from apps.export.models import ShipmentStatusLog

    ws = wb.active
    ws.title = 'Audit log'
    row = _write_title(ws, 'Audit log — soňky özgerişler', _period_label(from_date, to_date))
    row = _write_header(ws, row, ['Wagty', 'Ulanyjy', 'Maşyn kody', 'Ýagdaý', 'Bellik'])

    qs = (
        ShipmentStatusLog.objects
        .filter(changed_at__date__gte=from_date, changed_at__date__lte=to_date)
        .select_related('shipment', 'changed_by', 'status')
        .order_by('-changed_at')[:500]
    )
    for log in qs:
        ws.cell(row=row, column=1, value=log.changed_at.strftime('%Y-%m-%d %H:%M'))
        user = log.changed_by
        ws.cell(row=row, column=2, value=(user.get_full_name() or user.username) if user else '')
        ws.cell(row=row, column=3, value=log.shipment.cargo_code if log.shipment_id else '')
        ws.cell(row=row, column=4, value=log.status.name_tk if log.status_id else '')
        ws.cell(row=row, column=5, value=log.comment or '')
        row += 1
    _autosize(ws)


# ---------------------------------------------------------------------------
# Public dispatch
# ---------------------------------------------------------------------------

_SECTIONS = {
    'monthly':         _write_monthly,
    'firms':           _write_firms,
    'routes':          _write_routes,
    'blocks':          _write_blocks,
    'seasons_compare': _write_seasons,
    'audit':           _write_audit,
}


def build_excel(section: str, from_date: date, to_date: date) -> bytes:
    """Return a .xlsx workbook for the requested section as raw bytes.

    Args:
        section:   One of 'monthly' | 'firms' | 'routes' | 'blocks' |
                   'seasons_compare' | 'audit'.
        from_date: Inclusive start of the period.
        to_date:   Inclusive end of the period.

    Raises:
        ValueError: If the section name is not recognised.
    """
    writer = _SECTIONS.get(section)
    if writer is None:
        raise ValueError(f'Unknown export section: {section!r}')

    wb = Workbook()
    writer(wb, from_date, to_date)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
