"""Import weight/invoice details from Export_contracts_20252026_1.xlsx → Shipment (enrich).

Sources:
  - sheet '2-Sales': weight_net, truck plate, passport ref, scan status, R15 notes
  - sheet 'gross net': gross_kg, net_kg, box_count, pallet_count (right-side cols 7-11)

Join strategy:
  - gross_net rows are indexed by sequential serial (col 0) — build a dict keyed by int serial
  - 2-Sales rows each have a global serial (col 0) — look up gross_net data by that serial
  - Shipment match: (invoice_date, export_firm_code) from 2-Sales → DB lookup
    - If unique match: enrich the shipment
    - If multiple matches: match positionally within the group by serial order vs cargo_code sort
    - If no match: log warning, skip enrichment (comment still stored if has R15 note)

Enrichment written to Shipment:
  - weight_net (from col J) — only if <= 50,000
  - weight_gross (from gross_net right side col 7)
  - box_count (from gross_net right side col 9)
  - pallet_count (from gross_net right side col 10)
  - vehicle_responsible (truck plate from col L)
  - passport ref from col M stored in notes field

ShipmentComment created for:
  - Col O (R15 notes): prefix [Migrated from R15]
  - Cancelled rows (col N = yatyryldy/iptal/YZA SUYSIRILDI): note with cancel reason
"""
import datetime
import logging
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).parents[6] / 'data' / 'p3-export' / 'Export_contracts_20252026_1.xlsx'

CANCEL_VALUES = {'yatyryldy', 'iptal', 'YZA SUYSIRILDI', 'yza suysirildi'}

# All fields we may update on Shipment
UPDATE_FIELDS = [
    'weight_net', 'weight_gross', 'box_count', 'pallet_count',
    'vehicle_responsible', 'notes',
]


def _to_decimal(val):
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                return datetime.datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    return None


class Command(BaseCommand):
    help = 'Enrich Shipment records with weight/invoice details from Export_contracts_20252026_1.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?', default=str(DEFAULT_PATH))
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        from apps.export.models import Shipment, ShipmentComment
        from apps.core.models import ExportFirm, Customer, User

        path = Path(options['file'])
        if not path.exists():
            self.stderr.write(f'File not found: {path}')
            return

        dry_run = options['dry_run']

        # Use the first superuser as comment author for migrated notes
        system_user = User.objects.filter(is_superuser=True).first() or User.objects.first()
        if not system_user:
            self.stderr.write('No users in DB — cannot create ShipmentComment records.')
            return

        # --- Load gross_net data (serial -> dict) ---
        self.stdout.write('Loading gross_net sheet...')
        wb = openpyxl.load_workbook(str(path), data_only=True)
        gross_net_map = {}
        ws_gn = wb['gross net']
        for row in ws_gn.iter_rows(min_row=2, values_only=True):
            pad = list(row)
            while len(pad) < 12:
                pad.append(None)
            serial = pad[0]
            if serial is None:
                continue
            try:
                serial_int = int(serial)
            except (TypeError, ValueError):
                continue
            # Right-side columns (indices 7-11): BRUT, NET, YASIK, PALET, PALET_AGRAMY
            gross_net_map[serial_int] = {
                'weight_gross': _to_decimal(pad[7]),
                'box_count': int(pad[9]) if pad[9] is not None else None,
                'pallet_count': int(pad[10]) if pad[10] is not None else None,
            }

        self.stdout.write(f'  gross_net rows loaded: {len(gross_net_map)}')

        # --- Build shipment index keyed by (date, firm_code_upper) ---
        # For each key, shipments are sorted by cargo_code to enable positional matching.
        shipment_by_key = {}
        all_shipments = list(
            Shipment.objects.prefetch_related('firm_splits__export_firm')
        )
        for s in all_shipments:
            for split in s.firm_splits.all():
                key = (s.date, split.export_firm.code.upper())
                shipment_by_key.setdefault(key, []).append(s)
        for key in shipment_by_key:
            shipment_by_key[key].sort(key=lambda s: s.cargo_code)

        # --- Parse 2-Sales sheet ---
        self.stdout.write('Parsing 2-Sales sheet...')
        ws_s = wb['2-Sales']

        group_counters = {}  # (date, seller_upper) -> int (rows seen so far for this group)
        updates = []         # list of (Shipment, field_values_dict)
        comments = []        # list of ShipmentComment instances
        skipped_large = 0
        skipped_no_match = 0
        matched = 0
        warnings = []

        for row in ws_s.iter_rows(min_row=2, values_only=True):
            pad = list(row)
            while len(pad) < 16:
                pad.append(None)

            global_serial = pad[0]
            if global_serial is None:
                continue
            try:
                global_serial_int = int(global_serial)
            except (TypeError, ValueError):
                continue

            seller_raw = str(pad[1]).strip() if pad[1] else ''
            invoice_date = _parse_date(pad[4])
            weight_raw = pad[9]
            plate_raw = str(pad[11]).strip() if pad[11] else None
            passport_raw = str(pad[12]).strip() if pad[12] else None
            scan_val = pad[13]
            r15_note = (
                str(pad[14]).strip()
                if pad[14] and str(pad[14]).strip() not in ('', '-', 'None')
                else None
            )

            # Skip large batch entries
            if weight_raw and isinstance(weight_raw, (int, float)) and float(weight_raw) > 50000:
                skipped_large += 1
                warnings.append(
                    f'Row {global_serial_int}: weight={weight_raw} > 50,000 kg — skipped (batch entry)'
                )
                continue

            weight_net = _to_decimal(weight_raw)

            # Gross/net data from gross_net sheet
            gn = gross_net_map.get(global_serial_int, {})

            # Cancel detection
            is_cancelled = isinstance(scan_val, str) and scan_val.strip() in CANCEL_VALUES
            cancel_reason = scan_val.strip() if is_cancelled else None

            # Positional matching within (date, seller) group
            seller_upper = seller_raw.upper()
            group_key = (invoice_date, seller_upper)
            group_idx = group_counters.get(group_key, 0)
            group_counters[group_key] = group_idx + 1

            shipment = None
            candidates = shipment_by_key.get(group_key, [])
            if candidates:
                if group_idx < len(candidates):
                    shipment = candidates[group_idx]
                else:
                    warnings.append(
                        f'Row {global_serial_int}: group ({invoice_date}, {seller_raw}) '
                        f'index {group_idx} exceeds {len(candidates)} shipments — unmatched'
                    )

            # If no shipment and nothing to record, skip
            if shipment is None and not is_cancelled and r15_note is None:
                skipped_no_match += 1
                continue

            if shipment is not None:
                matched += 1

                # Collect field values for this shipment
                fv = {}
                if weight_net is not None:
                    fv['weight_net'] = weight_net
                if gn.get('weight_gross') is not None:
                    fv['weight_gross'] = gn['weight_gross']
                if gn.get('box_count') is not None:
                    fv['box_count'] = gn['box_count']
                if gn.get('pallet_count') is not None:
                    fv['pallet_count'] = gn['pallet_count']
                if plate_raw:
                    fv['vehicle_responsible'] = plate_raw
                if passport_raw:
                    existing = shipment.notes or ''
                    if passport_raw not in existing:
                        fv['notes'] = f'{existing}\nPassport: {passport_raw}'.strip()

                if fv:
                    # Apply immediately to the object so multiple enrichments per shipment accumulate
                    for field, val in fv.items():
                        setattr(shipment, field, val)
                    updates.append(shipment)

                if is_cancelled:
                    comments.append(ShipmentComment(
                        shipment=shipment,
                        user=system_user,
                        content=f'[Cancelled] {cancel_reason}',
                        is_system=True,
                    ))

                if r15_note:
                    comments.append(ShipmentComment(
                        shipment=shipment,
                        user=system_user,
                        content=f'[Migrated from R15] {r15_note}',
                        is_system=True,
                    ))
            else:
                if r15_note or is_cancelled:
                    warnings.append(
                        f'Row {global_serial_int}: seller={seller_raw}, date={invoice_date} '
                        f'— no matching shipment, note preserved: {r15_note or cancel_reason}'
                    )
                skipped_no_match += 1

        wb.close()

        for w in warnings:
            self.stderr.write(f'WARNING: {w}')

        if dry_run:
            self.stdout.write(
                f'[dry-run] Would update {len(updates)} shipments | '
                f'Create {len(comments)} comments | '
                f'Skipped large: {skipped_large} | Skipped no-match: {skipped_no_match} | '
                f'Matched: {matched} | Warnings: {len(warnings)}'
            )
            return

        updated_count = 0
        comment_count = 0

        with transaction.atomic():
            # Deduplicate updates — same shipment may appear multiple times if it had 2 rows
            seen_ids = set()
            unique_updates = []
            for s in updates:
                if s.pk not in seen_ids:
                    seen_ids.add(s.pk)
                    unique_updates.append(s)

            for i in range(0, len(unique_updates), 500):
                batch = unique_updates[i:i + 500]
                Shipment.objects.bulk_update(batch, UPDATE_FIELDS, batch_size=500)
                updated_count += len(batch)

            for i in range(0, len(comments), 500):
                batch = comments[i:i + 500]
                ShipmentComment.objects.bulk_create(batch, batch_size=500, ignore_conflicts=True)
                comment_count += len(batch)

        self.stdout.write(self.style.SUCCESS(
            f'Imported: {updated_count} | Skipped: {skipped_no_match + skipped_large} | '
            f'Comments created: {comment_count} | Warnings: {len(warnings)}'
        ))
