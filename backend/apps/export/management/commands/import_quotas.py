"""Import government export quotas (CYKAN KWOTA) from quota.xlsx into QuotaIssuance.

Reads sheet Kwota-2, "issued quota" section:
  - Tomato: firm headers row 8 cols B-P (15 firms), date col A, data rows 9-25.
  - Pepper: firm headers row 8 cols Y-Z (2 firms), date col X, data rows 9-10.
Groups per-firm amounts by issue date into QuotaIssuance + QuotaIssuanceFirmAllocation.

Idempotent: deletes existing issuances for each imported (issue_date, product_type)
before recreating. Issuances from a previous file with dates not in this file are
left untouched and reported.

Usage:
    python manage.py import_quotas                  # dry-run (default)
    python manage.py import_quotas --commit         # write to DB
    python manage.py import_quotas /path/to/file    # custom path
"""
import logging
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.export.models import QuotaIssuance, QuotaIssuanceFirmAllocation

from ._quota_import_utils import parse_quota_date, resolve_firm

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).parents[5] / 'data' / 'quota' / 'quota.xlsx'
SHEET_QUOTA = 'Kwota-2'

HEADER_ROW = 8

# Tomato: 15 firm columns B(2)-P(16), date in col A(1), issued data rows 9-25.
TOMATO_FIRM_COLS = list(range(2, 17))
TOMATO_DATE_COL = 1
TOMATO_DATA_ROWS = list(range(9, 26))

# Pepper: 2 firm columns Y(25)-Z(26), date in col X(24), issued data rows 9-10.
PEPPER_FIRM_COLS = [25, 26]
PEPPER_DATE_COL = 24
PEPPER_DATA_ROWS = [9, 10]


class Command(BaseCommand):
    help = 'Import issued quota (CYKAN KWOTA) from quota.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('path', nargs='?', default=str(DEFAULT_PATH))
        parser.add_argument('--commit', action='store_true', help='Write to DB (default: dry-run)')

    def handle(self, *args, **options):
        path = Path(options['path'])
        commit = options['commit']

        if not path.exists():
            self.stderr.write(f'File not found: {path}')
            return

        self.stdout.write(f'Loading {path} ...')
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[SHEET_QUOTA]
        firm_cache: dict = {}

        # {(issue_date, product_type): [(firm_name, kg)]}
        issuance_data: dict[tuple[date, str], list[tuple[str, Decimal]]] = defaultdict(list)

        def _read_section(firm_cols, date_col, data_rows, product):
            col_firms = {
                col: str(ws.cell(row=HEADER_ROW, column=col).value).strip()
                for col in firm_cols
                if ws.cell(row=HEADER_ROW, column=col).value
            }
            self.stdout.write(f'  {product.title()} section: {len(col_firms)} firms')

            for row_idx in data_rows:
                issue_date = parse_quota_date(
                    ws.cell(row=row_idx, column=date_col).value, fix_out_of_season=True,
                )
                if not issue_date:
                    continue
                for col, firm_name in col_firms.items():
                    val = ws.cell(row=row_idx, column=col).value
                    if not val:
                        continue
                    try:
                        kg = Decimal(str(val))
                    except (ValueError, ArithmeticError):
                        continue
                    if kg <= 0:
                        continue
                    issuance_data[(issue_date, product)].append((firm_name, kg))

        _read_section(TOMATO_FIRM_COLS, TOMATO_DATE_COL, TOMATO_DATA_ROWS, 'tomato')
        _read_section(PEPPER_FIRM_COLS, PEPPER_DATE_COL, PEPPER_DATA_ROWS, 'pepper')

        total_allocs = sum(len(v) for v in issuance_data.values())
        self.stdout.write(f'\n  Issuance events: {len(issuance_data)}')
        self.stdout.write(f'  Total firm allocations: {total_allocs}')
        for (d, p), allocs in sorted(issuance_data.items()):
            total = sum(kg for _, kg in allocs)
            self.stdout.write(f'    {d} | {p:6} | {len(allocs):2} firms | {total:>12,.0f} kg')

        # Report issuances already in the DB whose (date, product) is NOT in this
        # file — these were imported from an older file and will remain untouched.
        existing = set(
            QuotaIssuance.objects.values_list('issue_date', 'product_type')
        )
        incoming = set(issuance_data.keys())
        orphans = existing - incoming
        if orphans:
            self.stdout.write(self.style.WARNING(
                f'\n  {len(orphans)} existing issuance(s) NOT in this file '
                f'(will remain untouched): '
                + ', '.join(f'{d}/{p}' for d, p in sorted(orphans))
            ))

        if not commit:
            self.stdout.write('\n  DRY RUN — use --commit to write to DB')
            return

        with transaction.atomic():
            deleted = 0
            for (issue_date, product_type) in issuance_data:
                d, _ = QuotaIssuance.objects.filter(
                    issue_date=issue_date, product_type=product_type,
                ).delete()
                deleted += d
            if deleted:
                self.stdout.write(f'  Deleted {deleted} rows for re-imported (date, product) combos')

            created_issuances = created_allocs = skipped = 0
            for (issue_date, product_type), allocs in sorted(issuance_data.items()):
                iso = issue_date.isocalendar()
                issuance = QuotaIssuance.objects.create(
                    issue_date=issue_date,
                    product_type=product_type,
                    matched_week=iso[1],
                    matched_year=iso[0],
                    notes='Imported from quota.xlsx',
                )
                created_issuances += 1

                alloc_objs = []
                for firm_name, kg in allocs:
                    firm = resolve_firm(firm_name, firm_cache)
                    if firm is None:
                        skipped += 1
                        continue
                    alloc_objs.append(QuotaIssuanceFirmAllocation(
                        issuance=issuance, export_firm=firm, kg_quota=kg,
                    ))
                QuotaIssuanceFirmAllocation.objects.bulk_create(alloc_objs, batch_size=500)
                created_allocs += len(alloc_objs)

            self.stdout.write(self.style.SUCCESS(
                f'\n  Done: {created_issuances} issuances, {created_allocs} firm allocations, '
                f'{len(firm_cache)} firms resolved, {skipped} allocations skipped (firm not found).'
            ))
