"""Management command: import historical shipment data from Excel source files.

Sources:
  - Hasabat_202526.xlsx  → Saher sheet (primary: cargo codes, geography, firms)
  - Export_contracts_20252026_1.xlsx → 2-Sales sheet (secondary: weight_net, USD amount)
  - Export_contracts_20252026_1.xlsx → gross net sheet (secondary: weight_gross, box/pallet)

Architecture:
  Step 1: Read Saher → cargo_code → shipment data map
  Step 2: Read 2-Sales → (invoice_date, buyer_normalized) → [(seq_no, qty_kg, usd)] queue
  Step 3: Read gross net → seq_no → (brut_kg, net_kg, box_count, pallet_count, pallet_tare)
  Step 4: For each Saher row:
    a. Normalize and validate cargo code
    b. Resolve FK lookups (country, customer, import firm, export firms)
    c. Enrich with weight data from 2-Sales match (best-effort by date+buyer)
    d. Enrich with gross/packaging data from gross net (via matched 2-Sales seq_no)
    e. Create Shipment + ShipmentFirmSplit(s) + ShipmentComment(s)
  Step 5: Print summary report

Usage:
    python manage.py import_shipments
    python manage.py import_shipments --dry-run
    python manage.py import_shipments --dry-run --excel-dir /path/to/data
    python manage.py import_shipments --skip-existing
"""
import datetime
import logging
import os
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.test.utils import CaptureQueriesContext

from apps.core.models import (
    City,
    Country,
    Customer,
    ExportFirm,
    ImportFirm,
    Season,
    ShipmentStatusType,
    User,
)
from apps.export.models import Shipment, ShipmentComment, ShipmentFirmSplit

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────

BATCH_SIZE = 500

# Cargo code: DDCC###/YY  (day + 2-letter month abbrev + 3-digit seq + slash + 2-digit year)
CARGO_CODE_RE = re.compile(r'^\d{2}[A-Z]{2}\d{3}/\d{2}$')

# Month abbreviation → month number (all uppercase)
MONTH_ABBREV = {
    'SP': 9,   # September
    'OC': 10,  # October
    'NV': 11,  # November
    'DC': 12,  # December
    'JA': 1,   # January
    'FB': 2,   # February
    'MR': 3,   # March
    'AP': 4,   # April (may appear)
    'MY': 5,   # May
    'JN': 6,   # June
    'JL': 7,   # July
    'AG': 8,   # August
}

# Country label variants in Saher → DB Country.code
COUNTRY_LABEL_TO_CODE = {
    'GAZAGYSTAN': 'KZ',
    'GAZAKSTAN': 'KZ',
    'KAZAKSTAN': 'KZ',
    'KZ': 'KZ',
    'ROSSIYA': 'RU',
    'RUSSIA': 'RU',
    'RU': 'RU',
    'OZBEKYSTAN': 'UZ',
    'OZBEKISTAN': 'UZ',
    'UZ': 'UZ',
    'GYRGYZYSTAN': 'KG',
    'GYRGYZSTAN': 'KG',
    'KG': 'KG',
    'TAJIGISTAN': 'TJ',
    'TJ': 'TJ',
    'BELARUS': 'BY',
    'BY': 'BY',
    'OWGANYSTAN': 'AF',
    'AF': 'AF',
}

# Saher firm name variants → ExportFirm.code
# Full names, partial matches, and known alias patterns
FIRM_NAME_TO_CODE = {
    # Full name variants
    'GOKBULUT H.J': 'GB',
    'GOK BULUT HJ': 'GB',
    'GOKBULUT HJ': 'GB',
    'HEMSAYA H.J': 'HMS',
    'HEMSAYA HJ': 'HMS',
    '"HEMSAYA" HJ': 'HMS',
    'MIWELI ATYZ H.J': 'MA',
    'MIWELI ATYZ HJ': 'MA',
    '"MIWELI ATYZ" HJ': 'MA',
    'DATLY MIWE H.J': 'DM',
    'DATLY MIWE HJ': 'DM',
    '"DATLY MIWE" HJ': 'DM',
    'YIGIT H.J': 'YE',
    'YIGIT HJ': 'YE',
    '"YIGIT" HJ': 'YE',
    'YGTYBARLY ENJAM': 'YE',
    'YGTYBARLY ENJAMLAR': 'YE',
    '"YGTYBARLY ENJAMLAR" JH': 'YE',
    'AK BULUT HJ': 'AB',
    '"AK BULUT" HJ': 'AB',
    'ISGAR HJ': 'ISG',
    '"ISGAR" HJ': 'ISG',
    'YUMAK H.J': 'YMK',
    'YUMAK HJ': 'YMK',
    '"YUMAK" HJ': 'YMK',
    'KERWENLI ILLER HK': 'KIH',
    # Short codes that appear as-is
    'YGT': 'YGT',
    'HMS': 'HMS',
    'YE': 'YE',
    'DM': 'DM',
    'GB': 'GB',
    'MA': 'MA',
    'AB': 'AB',
    'ISG': 'ISG',
    'YMK': 'YMK',
    # Tel (individual entrepreneurs) — map to their codes
    'TEL DOWRANOW E': 'Tel ED',
    'TEL DOWRANOW J': 'Tel JD',
    'TEL HEMIDOW P': 'Tel PH',
    'TEL HEMIDOW C': 'Tel CH',
    'TEL AMANGELDIYEW G': 'Tel GA',
    'TEL HEMIDOW J': 'Tel JD',
}

# Cancelled status values in 2-Sales col N
CANCELLED_VALUES = {'yatyryldy', 'iptal', 'YZA SUYSIRILDI'}

# Weight threshold above which a row is flagged as a multi-truck batch outlier
BATCH_WEIGHT_THRESHOLD_KG = 50_000


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_cargo_code(raw: str) -> str:
    """Replace Cyrillic С (U+0421) with Latin C (U+0043) and strip whitespace."""
    return raw.strip().replace('\u0421', 'C')


def _validate_cargo_code(code: str) -> bool:
    """Return True if code matches DDCC###/YY after normalization."""
    return bool(CARGO_CODE_RE.match(code))


def _parse_date_from_cargo_code(code: str) -> datetime.date | None:
    """Derive the shipment date from a normalized cargo code.

    Format: DDCC###/YY
    Returns None if month abbreviation is unknown or date is invalid.
    """
    if len(code) < 10:
        return None
    dd_str = code[:2]
    cc = code[2:4].upper()
    yy_str = code[8:10]
    month = MONTH_ABBREV.get(cc)
    if month is None:
        return None
    try:
        dd = int(dd_str)
        yy = int(yy_str) + 2000
        return datetime.date(yy, month, dd)
    except (ValueError, OverflowError):
        return None


def _parse_firm_codes(raw_firm: str) -> list[str]:
    """Parse a Saher firm field into a list of ExportFirm codes.

    Handles:
    - Single full name: 'Datly Miwe H.J' → ['DM']
    - Hyphenated code pair: 'YGT-HMS' → ['YGT', 'HMS']
    - Plus-separated pair: 'Tel Hem P+YE' → ['Tel PH', 'YE']
    - Slash-separated pair: 'Tel Dowranow J/HMS' → ['Tel JD', 'HMS']
    - Triple split: 'YGT-HemP-YE' → ['YGT', 'Tel PH', 'YE']
    - Special weighted: 'Ygt 14 tonna  Oguz Yoly 3 tonna' → ['YGT', 'YMK'] (best-effort)
    """
    if not raw_firm:
        return []

    raw = raw_firm.strip()

    # Special case: weighted description with 'tonna'
    if 'tonna' in raw.lower():
        codes = []
        # 'Ygt 14 tonna  Oguz Yoly 3 tonna' → YGT + YMK
        lower = raw.lower()
        if 'ygt' in lower or 'ygit' in lower:
            codes.append('YGT')
        if 'oguz' in lower or 'yoly' in lower:
            codes.append('YMK')
        return codes if codes else [_lookup_firm_code(raw)]

    # Try simple lookup first (full name or short code)
    direct = _lookup_firm_code(raw)
    if direct:
        return [direct]

    # Split on separator characters: -, +, /
    # Order matters: try '+' and '/' first, then '-' (since '-' is also in single firm names)
    for sep in ['+', '/']:
        if sep in raw:
            parts = [p.strip() for p in raw.split(sep)]
            codes = [_lookup_firm_code(p) for p in parts]
            resolved = [c for c in codes if c]
            if resolved:
                return resolved

    # Hyphen splitting: must be CODE-CODE pattern (all-caps short codes with '-')
    # Pattern: each part is either a short code (≤6 chars, no spaces) or known alias
    if '-' in raw:
        parts = [p.strip() for p in raw.split('-')]
        # Only treat as multi-firm if all parts look like codes (no spaces in parts)
        if all(len(p) <= 10 and ' ' not in p for p in parts):
            codes = [_lookup_firm_code(p) for p in parts]
            resolved = [c for c in codes if c]
            if len(resolved) == len(parts):
                return resolved

    # Fallback: try to resolve as single entity
    code = _lookup_firm_code(raw)
    return [code] if code else []


def _lookup_firm_code(name: str) -> str | None:
    """Look up a firm code from a name string. Returns None if unknown."""
    if not name:
        return None
    # Normalize: strip, upper
    normalized = name.strip().upper()
    # Direct map lookup
    if normalized in FIRM_NAME_TO_CODE:
        return FIRM_NAME_TO_CODE[normalized]
    # Try prefix matching for 'Tel ...' names
    tel_map = {
        'TEL DOWRANOW E': 'Tel ED',
        'TEL DOWRANOW J': 'Tel JD',
        'TEL HEMIDOW P': 'Tel PH',
        'TEL HEMIDOW C': 'Tel CH',
        'TEL AMANGELDIYEW G': 'Tel GA',
        'TEL HEM P': 'Tel PH',
        'TEL HEM C': 'Tel CH',
    }
    for prefix, code in tel_map.items():
        if normalized.startswith(prefix):
            return code
    # Short known codes (2–6 chars, all letters/digits)
    if re.match(r'^[A-Z]{2,6}$', normalized):
        return normalized  # return as-is; DB lookup will validate
    return None


def _safe_decimal(value) -> Decimal | None:
    """Convert a value to Decimal or return None if not convertible."""
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _normalize_str(value) -> str:
    """Strip whitespace from a string cell value."""
    if value is None:
        return ''
    return str(value).strip()


# ── Excel readers ──────────────────────────────────────────────────────────────

def _read_saher_sheet(wb_path: str) -> list[dict]:
    """Read Hasabat_202526.xlsx Saher sheet into a list of dicts.

    Columns (0-indexed, data starts at row 4):
      0: cargo_code (Ýük Kody)
      1: block_label (Pomidoryň ýygylan bölümi)
      2: export_firm_raw (Eksport eden Firma)
      3: country_raw (Eksport ýurdy)
      4: customer_name (Müşderi ady / telefon no)
      5: city_name (Şäheri)
      6: import_firm_name (Import Edilen Firma)
      7: status_note (Ýygym ýagdaýy / harvest status)
    """
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['Saher']
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        raw_code = row[0]
        if not isinstance(raw_code, str):
            continue
        rows.append({
            'cargo_code_raw': raw_code,
            'block_label': _normalize_str(row[1] if len(row) > 1 else None),
            'export_firm_raw': _normalize_str(row[2] if len(row) > 2 else None),
            'country_raw': _normalize_str(row[3] if len(row) > 3 else None),
            'customer_name': _normalize_str(row[4] if len(row) > 4 else None),
            'city_name': _normalize_str(row[5] if len(row) > 5 else None),
            'import_firm_name': _normalize_str(row[6] if len(row) > 6 else None),
            'status_note': _normalize_str(row[7] if len(row) > 7 else None),
        })
    wb.close()
    return rows


def _read_sales_sheet(wb_path: str) -> dict:
    """Read Export_contracts 2-Sales sheet.

    Returns a dict:
      {
        'by_date_buyer': defaultdict(list),  # key=(date, buyer_norm) → [(seq_no, qty_kg, usd, r15, cancelled)]
        'seq_to_row': dict,                  # seq_no → full row dict
      }

    Columns (0-indexed, header row 1, data from row 2):
      0: seq_no (global sequential number)
      1: seller_code
      2: buyer_name
      3: contract_number
      4: invoice_date (datetime)
      5: total_trucks_in_contract
      6: serial_no_within_contract
      7: invoice_number
      8: incoterms
      9: qty_kg (weight_net)
      10: usd_amount
      11: truck_plate
      12: passport_deal_ref
      13: scan_status (True/False/'-'/'yatyryldy'/'iptal'/'YZA SUYSIRILDI')
      14: r15_note (free text)
    """
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['2-Sales']

    by_date_buyer = defaultdict(list)
    seq_to_row = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 10 or row[0] is None:
            continue
        try:
            seq_no = int(row[0])
        except (ValueError, TypeError):
            continue

        raw_date = row[4]
        if raw_date is None:
            continue
        if isinstance(raw_date, datetime.datetime):
            invoice_date = raw_date.date()
        elif isinstance(raw_date, datetime.date):
            invoice_date = raw_date
        else:
            continue  # skip unparseable dates

        buyer_raw = _normalize_str(row[2])
        buyer_norm = buyer_raw.upper()
        qty_kg = _safe_decimal(row[9])
        usd = _safe_decimal(row[10])
        r15_note = _normalize_str(row[14] if len(row) > 14 else None)
        scan_raw = row[13] if len(row) > 13 else None
        invoice_no = _normalize_str(row[7])

        cancelled_label = None
        if isinstance(scan_raw, str) and scan_raw.strip() in CANCELLED_VALUES:
            cancelled_label = scan_raw.strip()

        entry = {
            'seq_no': seq_no,
            'seller_code': _normalize_str(row[1]),
            'buyer_raw': buyer_raw,
            'buyer_norm': buyer_norm,
            'invoice_date': invoice_date,
            'invoice_no': invoice_no,
            'qty_kg': qty_kg,
            'usd': usd,
            'truck_plate': _normalize_str(row[11] if len(row) > 11 else None),
            'r15_note': r15_note,
            'cancelled_label': cancelled_label,
        }
        seq_to_row[seq_no] = entry
        key = (invoice_date, buyer_norm)
        by_date_buyer[key].append(entry)

    wb.close()
    return {'by_date_buyer': by_date_buyer, 'seq_to_row': seq_to_row}


def _read_gross_net_sheet(wb_path: str) -> dict[int, dict]:
    """Read Export_contracts gross net sheet.

    Returns: seq_no → {brut_kg, net_kg, box_count, pallet_count, pallet_tare}

    Columns (0-indexed, header row 1, data from row 2):
      0: seq_no
      1: brut_kg  (gross weight)
      2: net_kg   (net weight)
      3: yasik    (box count)
      4: palet    (pallet count)
      5: palet_agr (pallet tare weight kg)
    """
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['gross net']

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3 or row[0] is None:
            continue
        try:
            seq_no = int(row[0])
        except (ValueError, TypeError):
            continue

        brut = _safe_decimal(row[1] if len(row) > 1 else None)
        net = _safe_decimal(row[2] if len(row) > 2 else None)
        box = row[3] if len(row) > 3 else None
        palet = row[4] if len(row) > 4 else None
        palet_tare = _safe_decimal(row[5] if len(row) > 5 else None)

        gross_anomaly = False
        if brut is not None and net is not None and brut < net:
            gross_anomaly = True

        try:
            box_int = int(box) if box is not None else None
        except (ValueError, TypeError):
            box_int = None
        try:
            palet_int = int(palet) if palet is not None else None
        except (ValueError, TypeError):
            palet_int = None

        result[seq_no] = {
            'brut_kg': None if gross_anomaly else brut,
            'net_kg': net,
            'box_count': box_int,
            'pallet_count': palet_int,
            'pallet_tare': palet_tare,
            'gross_anomaly': gross_anomaly,
        }

    wb.close()
    return result


# ── Reference data cache ───────────────────────────────────────────────────────

class _ReferenceCache:
    """Loads and caches core reference objects for fast lookup during import."""

    def __init__(self, stdout):
        self._stdout = stdout
        self._countries: dict[str, Country] = {}       # code → Country
        self._customers: dict[str, Customer] = {}      # name_upper → Customer
        self._import_firms: dict[str, ImportFirm] = {} # name_upper → ImportFirm
        self._export_firms: dict[str, ExportFirm] = {} # code → ExportFirm
        self._cities: dict[tuple, City] = {}           # (country_code, name_upper) → City
        self._season: Season | None = None
        self._tamamlandy: ShipmentStatusType | None = None
        self._admin_user: User | None = None

        self._created_customers = 0
        self._created_import_firms = 0
        self._created_cities = 0

    def load(self) -> None:
        """Populate all caches from the database."""
        for c in Country.objects.all():
            self._countries[c.code] = c

        for c in Customer.objects.all():
            self._customers[c.name.upper()] = c

        for f in ImportFirm.objects.all():
            key = (f.name_tk or '').upper()
            self._import_firms[key] = f
            if f.name_en:
                self._import_firms[f.name_en.upper()] = f

        for f in ExportFirm.objects.all():
            self._export_firms[f.code] = f

        for city in City.objects.select_related('country').all():
            key = (city.country.code, city.name.upper())
            self._cities[key] = city

        self._season = Season.objects.filter(is_active=True).first()
        if not self._season:
            self._season = Season.objects.order_by('-start_date').first()

        self._tamamlandy = ShipmentStatusType.objects.filter(code='tamamlandy').first()

        self._admin_user = (
            User.objects.filter(is_superuser=True).order_by('id').first()
            or User.objects.order_by('id').first()
        )

        self._stdout.write(
            f'  Reference cache: {len(self._countries)} countries, '
            f'{len(self._export_firms)} export firms, '
            f'{len(self._import_firms)} import firms, '
            f'{len(self._customers)} customers'
        )

    def get_country(self, raw: str) -> Country | None:
        code = COUNTRY_LABEL_TO_CODE.get(raw.upper().strip())
        if code:
            return self._countries.get(code)
        return None

    def get_or_create_customer(self, name: str) -> Customer | None:
        if not name:
            return None
        key = name.upper()
        if key in self._customers:
            return self._customers[key]
        country = self._countries.get('KZ')  # default; overridden per-shipment
        obj = Customer.objects.create(name=name, default_country=country)
        self._customers[key] = obj
        self._created_customers += 1
        return obj

    def get_or_create_import_firm(self, name: str, country: Country | None) -> ImportFirm | None:
        if not name:
            return None
        key = name.upper()
        if key in self._import_firms:
            return self._import_firms[key]
        obj = ImportFirm.objects.create(name_tk=name, country=country)
        self._import_firms[key] = obj
        self._created_import_firms += 1
        return obj

    def get_export_firm(self, code: str) -> ExportFirm | None:
        return self._export_firms.get(code)

    def get_or_create_city(self, name: str, country: Country | None) -> City | None:
        if not name or country is None:
            return None
        normalized = name.strip()
        key = (country.code, normalized.upper())
        if key in self._cities:
            return self._cities[key]
        obj, _ = City.objects.get_or_create(
            country=country,
            name=normalized,
        )
        self._cities[key] = obj
        self._created_cities += 1
        return obj

    @property
    def season(self) -> Season | None:
        return self._season

    @property
    def tamamlandy(self) -> ShipmentStatusType | None:
        return self._tamamlandy

    @property
    def admin_user(self) -> User | None:
        return self._admin_user

    @property
    def stats(self) -> dict:
        return {
            'customers_created': self._created_customers,
            'import_firms_created': self._created_import_firms,
            'cities_created': self._created_cities,
        }


# ── Main command ───────────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = 'Import historical shipment data from Hasabat_202526.xlsx and Export_contracts_20252026_1.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Parse and validate all data but do not write to the database.',
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            default=False,
            help='Skip rows where a Shipment with matching cargo_code already exists.',
        )
        parser.add_argument(
            '--excel-dir',
            default=None,
            help='Directory containing the Excel files. Defaults to <project-root>/data/p3-export/',
        )

    def handle(self, *args, **options):
        dry_run = options['dry_run']
        skip_existing = options['skip_existing']
        excel_dir = options['excel_dir']

        # Resolve Excel directory
        if excel_dir is None:
            # Default: <project-root>/data/p3-export/
            backend_dir = os.path.dirname(  # backend/
                os.path.dirname(               # apps/
                    os.path.dirname(           # export/
                        os.path.dirname(       # management/
                            os.path.dirname(   # commands/
                                os.path.abspath(__file__)
                            )
                        )
                    )
                )
            )
            project_root = os.path.dirname(backend_dir)
            excel_dir = os.path.join(project_root, 'data', 'p3-export')

        hasabat_path = os.path.join(excel_dir, 'Hasabat_202526.xlsx')
        contracts_path = os.path.join(excel_dir, 'Export_contracts_20252026_1.xlsx')

        for path in (hasabat_path, contracts_path):
            if not os.path.exists(path):
                raise CommandError(f'Excel file not found: {path}')

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no data will be written.\n'))

        # ── Step 1–3: Load Excel data ──────────────────────────────────────────
        self.stdout.write('Reading Saher sheet...')
        saher_rows = _read_saher_sheet(hasabat_path)
        self.stdout.write(f'  {len(saher_rows)} rows read from Saher')

        self.stdout.write('Reading 2-Sales sheet...')
        sales_data = _read_sales_sheet(contracts_path)
        sales_lookup = sales_data['by_date_buyer']
        self.stdout.write(
            f'  {len(sales_data["seq_to_row"])} rows read from 2-Sales '
            f'({len(sales_lookup)} unique date+buyer keys)'
        )

        self.stdout.write('Reading gross net sheet...')
        gross_net = _read_gross_net_sheet(contracts_path)
        self.stdout.write(f'  {len(gross_net)} rows read from gross net')

        # ── Step 4: Load reference cache ──────────────────────────────────────
        self.stdout.write('Loading reference data...')
        cache = _ReferenceCache(self.stdout)
        cache.load()

        if cache.season is None:
            raise CommandError(
                'No Season found in database. Run: python manage.py seed_data'
            )
        if cache.tamamlandy is None:
            raise CommandError(
                'ShipmentStatusType "tamamlandy" not found. Run: python manage.py seed_data'
            )
        if cache.admin_user is None:
            raise CommandError(
                'No User found in database. Run: python manage.py seed_data'
            )

        # ── Step 5: Process rows ───────────────────────────────────────────────
        # Counters
        cnt_imported = 0
        cnt_skipped_existing = 0
        cnt_skipped_batch = 0
        cnt_skipped_no_code = 0
        cnt_skipped_invalid_code = 0
        cnt_sales_matched = 0
        cnt_gross_matched = 0
        cnt_r15_comments = 0
        cnt_cancelled_comments = 0
        cnt_firm_splits = 0
        cnt_missing_firms = 0
        warnings: list[str] = []
        errors: list[str] = []

        # Track cargo codes seen in this run to detect intra-file duplicates
        seen_codes: set[str] = set()

        # Mutable queues: when we match a Saher row to 2-Sales by date+buyer,
        # we pop from the front so each 2-Sales row is consumed at most once.
        # We work on a copy of the lists so the original stays intact for reporting.
        sales_queues: dict[tuple, list] = {k: list(v) for k, v in sales_lookup.items()}

        # Collect objects for bulk_create
        shipments_to_create: list[Shipment] = []
        # splits and comments keyed by cargo_code (filled after shipment creation)
        pending_splits: dict[str, list[dict]] = {}   # cargo_code → [{'firm_code':..., 'order':...}]
        pending_comments: dict[str, list[str]] = {}  # cargo_code → [content, ...]

        # ── Process each Saher row ─────────────────────────────────────────────
        for row_idx, row in enumerate(saher_rows, start=4):

            # 1. Normalize and validate cargo code
            raw_code = row['cargo_code_raw']
            normalized = _normalize_cargo_code(raw_code)

            if not normalized:
                cnt_skipped_no_code += 1
                continue

            if not _validate_cargo_code(normalized):
                cnt_skipped_invalid_code += 1
                warnings.append(
                    f'Row {row_idx}: invalid cargo code {raw_code!r} (normalized: {normalized!r}) — skipped'
                )
                continue

            if normalized in seen_codes:
                warnings.append(
                    f'Row {row_idx}: duplicate cargo code {normalized!r} within source file — skipped'
                )
                cnt_skipped_invalid_code += 1
                continue
            seen_codes.add(normalized)

            # 2. Skip if already exists
            if skip_existing and Shipment.objects.filter(cargo_code=normalized).exists():
                cnt_skipped_existing += 1
                continue

            # 3. Derive date from cargo code
            shipment_date = _parse_date_from_cargo_code(normalized)
            if shipment_date is None:
                warnings.append(
                    f'Row {row_idx}: cannot parse date from cargo code {normalized!r} — skipped'
                )
                cnt_skipped_invalid_code += 1
                continue

            # 4. Resolve country
            country_raw = row['country_raw']
            country = cache.get_country(country_raw) if country_raw else None
            if not country and country_raw:
                warnings.append(
                    f'Row {row_idx} ({normalized}): unknown country {country_raw!r}'
                )

            # 5. Resolve city (nullable — decided late)
            city_name = row['city_name']
            city = None
            if city_name and city_name.lower() not in ('gapy satyş', 'gapy satys', '-', ''):
                city = cache.get_or_create_city(city_name, country)

            # 6. Detect Gapy Satys flag
            is_gapy_satys = False
            if city_name and 'gapy' in city_name.lower():
                is_gapy_satys = True
            customer_raw = row['customer_name']
            if customer_raw and 'gapy' in customer_raw.lower():
                is_gapy_satys = True

            # 7. Resolve customer
            customer_name = customer_raw
            if customer_name and 'gapy' in customer_name.lower():
                customer_name = 'ÝGT Gapy Satyş'
            customer = cache.get_or_create_customer(customer_name) if customer_name else None

            # 8. Resolve import firm
            import_firm_name = row['import_firm_name']
            import_firm = (
                cache.get_or_create_import_firm(import_firm_name, country)
                if import_firm_name else None
            )

            # 9. Parse export firm codes
            firm_codes = _parse_firm_codes(row['export_firm_raw'])

            # 10. Try to match to 2-Sales for weight enrichment
            #     Key: (date, buyer_normalized)
            #     Buyer in 2-Sales ≈ import firm name
            weight_net: Decimal | None = None
            weight_gross: Decimal | None = None
            box_count: int | None = None
            pallet_count: int | None = None
            total_usd: Decimal | None = None
            matched_seq_no: int | None = None

            if import_firm_name and shipment_date:
                buyer_key = import_firm_name.strip().upper()
                lookup_key = (shipment_date, buyer_key)
                queue = sales_queues.get(lookup_key)
                if queue:
                    sales_entry = queue.pop(0)
                    matched_seq_no = sales_entry['seq_no']
                    weight_net = sales_entry['qty_kg']
                    total_usd = sales_entry['usd']
                    cnt_sales_matched += 1

                    # R15 note from 2-Sales
                    if sales_entry['r15_note']:
                        r15_content = f'[Migrated from R15] {sales_entry["r15_note"]}'
                        pending_comments.setdefault(normalized, []).append(r15_content)
                        cnt_r15_comments += 1

                    # Cancelled label
                    if sales_entry['cancelled_label']:
                        cancel_content = f'[Cancelled: {sales_entry["cancelled_label"]}]'
                        pending_comments.setdefault(normalized, []).append(cancel_content)
                        cnt_cancelled_comments += 1

                    # 11. Enrich from gross net sheet using matched seq_no
                    gn = gross_net.get(matched_seq_no)
                    if gn:
                        cnt_gross_matched += 1
                        if gn['gross_anomaly']:
                            warnings.append(
                                f'Row {row_idx} ({normalized}): gross < net anomaly '
                                f'in gross_net seq {matched_seq_no} — weight_gross set to NULL'
                            )
                            weight_gross = None
                        else:
                            weight_gross = gn['brut_kg']
                        # Prefer gross_net net over 2-Sales qty when available
                        if gn['net_kg'] is not None:
                            weight_net = gn['net_kg']
                        box_count = gn['box_count']
                        pallet_count = gn['pallet_count']

            # 12. Validate weight reasonableness
            if weight_net is not None and weight_net > BATCH_WEIGHT_THRESHOLD_KG and not is_gapy_satys:
                warnings.append(
                    f'Row {row_idx} ({normalized}): weight_net={weight_net} kg exceeds '
                    f'{BATCH_WEIGHT_THRESHOLD_KG} kg batch threshold — skipped'
                )
                cnt_skipped_batch += 1
                continue

            if (weight_net is not None and weight_gross is not None
                    and weight_gross < weight_net):
                warnings.append(
                    f'Row {row_idx} ({normalized}): weight_gross ({weight_gross}) < '
                    f'weight_net ({weight_net}) — weight_gross set to NULL'
                )
                weight_gross = None

            # 13. Build Shipment object
            shipment = Shipment(
                cargo_code=normalized,
                date=shipment_date,
                season=cache.season,
                status=cache.tamamlandy,
                country=country,
                city=city,
                customer=customer,
                import_firm=import_firm,
                is_gapy_satys=is_gapy_satys,
                weight_net=weight_net,
                weight_gross=weight_gross,
                box_count=box_count,
                pallet_count=pallet_count,
                total_amount_usd=total_usd,
                created_by=cache.admin_user,
            )
            shipments_to_create.append(shipment)

            # 14. Queue firm splits
            if firm_codes:
                resolved_codes = []
                for code in firm_codes:
                    firm_obj = cache.get_export_firm(code)
                    if firm_obj:
                        resolved_codes.append(code)
                    else:
                        warnings.append(
                            f'Row {row_idx} ({normalized}): export firm code {code!r} '
                            f'not found in DB — split skipped'
                        )
                        cnt_missing_firms += 1
                if resolved_codes:
                    pending_splits[normalized] = [
                        {'firm_code': c, 'order': i + 1}
                        for i, c in enumerate(resolved_codes)
                    ]
            else:
                warnings.append(
                    f'Row {row_idx} ({normalized}): no export firm resolved from '
                    f'{row["export_firm_raw"]!r}'
                )

            cnt_imported += 1

            # Flush shipment batch at BATCH_SIZE to keep memory bounded
            if len(shipments_to_create) >= BATCH_SIZE and not dry_run:
                self._flush_shipments(
                    shipments_to_create, pending_splits, pending_comments,
                    cache, skip_existing, cnt_r15_comments,
                )
                shipments_to_create.clear()

        # ── Flush remaining ────────────────────────────────────────────────────
        if not dry_run and shipments_to_create:
            self._flush_shipments(
                shipments_to_create, pending_splits, pending_comments,
                cache, skip_existing, cnt_r15_comments,
            )
            shipments_to_create.clear()

        # ── Summary report ─────────────────────────────────────────────────────
        self.stdout.write('')
        self.stdout.write('=== Import Summary ===')
        if dry_run:
            self.stdout.write(self.style.WARNING('(DRY RUN — nothing was written)'))
        self.stdout.write(
            self.style.SUCCESS(f'Imported:               {cnt_imported:,} shipments')
        )
        self.stdout.write(f'Skipped (existing):     {cnt_skipped_existing:,}')
        self.stdout.write(f'Skipped (batch weight): {cnt_skipped_batch:,}')
        self.stdout.write(f'Skipped (bad code):     {cnt_skipped_invalid_code:,}')
        self.stdout.write(f'Skipped (no code):      {cnt_skipped_no_code:,}')
        self.stdout.write(f'2-Sales matched:        {cnt_sales_matched:,}')
        self.stdout.write(f'gross net matched:      {cnt_gross_matched:,}')
        self.stdout.write(f'R15 comments migrated:  {cnt_r15_comments:,}')
        self.stdout.write(f'Cancelled comments:     {cnt_cancelled_comments:,}')
        self.stdout.write(f'Firm splits created:    {cnt_firm_splits:,}')
        self.stdout.write(f'Missing firm refs:      {cnt_missing_firms:,}')
        self.stdout.write(
            f'Customers created:      {cache.stats["customers_created"]:,}'
        )
        self.stdout.write(
            f'Import firms created:   {cache.stats["import_firms_created"]:,}'
        )
        self.stdout.write(
            f'Cities created:         {cache.stats["cities_created"]:,}'
        )
        self.stdout.write(f'Warnings:               {len(warnings):,}')
        self.stdout.write(f'Errors:                 {len(errors):,}')

        if warnings:
            self.stdout.write('')
            self.stdout.write('--- Warnings (first 30) ---')
            for w in warnings[:30]:
                self.stdout.write(self.style.WARNING(f'  {w}'))
            if len(warnings) > 30:
                self.stdout.write(f'  ... and {len(warnings) - 30} more.')

        if errors:
            self.stdout.write('')
            self.stdout.write('--- Errors ---')
            for e in errors:
                self.stdout.write(self.style.ERROR(f'  {e}'))

    def _flush_shipments(
        self,
        shipments: list[Shipment],
        pending_splits: dict,
        pending_comments: dict,
        cache: _ReferenceCache,
        skip_existing: bool,
        r15_count: int,
    ) -> None:
        """Bulk-create a batch of Shipment objects, then create their splits and comments.

        Wraps each flush in a savepoint so a single bad batch doesn't abort everything.
        """
        if not shipments:
            return

        with transaction.atomic():
            # Filter out duplicates if --skip-existing is off (unique constraint safeguard)
            existing_codes = set(
                Shipment.objects.filter(
                    cargo_code__in=[s.cargo_code for s in shipments]
                ).values_list('cargo_code', flat=True)
            )
            to_insert = [s for s in shipments if s.cargo_code not in existing_codes]

            if not to_insert:
                return

            created = Shipment.objects.bulk_create(to_insert, batch_size=BATCH_SIZE)

            # Reload the created shipments to get their PKs
            code_to_shipment = {s.cargo_code: s for s in created}

            # Build firm splits
            splits_to_create = []
            for shipment in created:
                split_defs = pending_splits.pop(shipment.cargo_code, [])
                weight_per_firm = (
                    (shipment.weight_net / len(split_defs))
                    if shipment.weight_net and split_defs
                    else Decimal('0.00')
                )
                for split_def in split_defs:
                    firm_obj = cache.get_export_firm(split_def['firm_code'])
                    if firm_obj is None:
                        continue
                    splits_to_create.append(ShipmentFirmSplit(
                        shipment=shipment,
                        export_firm=firm_obj,
                        weight_kg=weight_per_firm,
                        amount_usd=None,
                        split_order=split_def['order'],
                    ))

            if splits_to_create:
                ShipmentFirmSplit.objects.bulk_create(
                    splits_to_create, batch_size=BATCH_SIZE
                )

            # Build comments
            comments_to_create = []
            for shipment in created:
                comment_contents = pending_comments.pop(shipment.cargo_code, [])
                for content in comment_contents:
                    comments_to_create.append(ShipmentComment(
                        shipment=shipment,
                        user=cache.admin_user,
                        content=content,
                        is_system=True,
                    ))

            if comments_to_create:
                ShipmentComment.objects.bulk_create(
                    comments_to_create, batch_size=BATCH_SIZE
                )
