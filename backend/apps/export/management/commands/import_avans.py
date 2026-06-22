"""Import the customs/document cash-advance ledger from data/avans.xlsx.

Source: data/avans.xlsx -> sheet 'Sheet1'

Structure (one running cash ledger, reconciles to zero):
  Col C  Senesi       - date
  Col D  Cykys Kody   - internal trip code (MY471, JN055, 15JN089/26) — NOT the
                        DB shipment_code, kept verbatim as export_code_raw
  Col E  Firma        - route firm pair / responsible person (HMS-DM, Tel Gurban J)
  Col F  Acyklama     - description -> drives category + label_raw
  Col G  Masyn nomeri - truck/trailer plate (4656AHF/2405TAH) — the shipment join key
  Col H  Alan pul     - money IN  -> FinansistAdvance (float top-up)
  Col I  Cykan pul    - money OUT -> CustomsExpense (customs/document fee)

Imports into:
  1. FinansistAdvance — one row per money-IN line (currency TMT, purpose = Acyklama)
  2. CustomsExpense   — one row per money-OUT line (category mapped from Acyklama)

Shipment linking: the Excel trip codes do not match Shipment.shipment_code, so each
expense is linked by a UNIQUE full-plate match of Masyn nomeri against
Shipment.truck_plate. Rows with no plate (batch fees), an unmatched plate, or an
ambiguous (multi-shipment) plate are imported with shipment=NULL — the verbatim
raw fields keep them fully auditable and they can be linked later once plates are
entered on more shipments.

Idempotency: every imported row carries the marker '[import:avans]' in its notes.
A (non-dry) run first DELETES all rows bearing that marker, then re-inserts — so
re-running is safe and reflects the current spreadsheet exactly.
"""
import datetime
import logging
import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.core.models import User
from apps.export.models import (
    CustomsExpense,
    CustomsExpenseCategory,
    FinansistAdvance,
    Shipment,
)

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).resolve().parents[5] / 'data' / 'avans.xlsx'
SHEET_NAME = 'Sheet1'
IMPORT_MARKER = '[import:avans]'
CURRENCY = 'TMT'

# Column indexes (1-based, openpyxl).
COL_DATE, COL_CODE, COL_FIRM, COL_DESC, COL_PLATE, COL_IN, COL_OUT = 3, 4, 5, 6, 7, 8, 9

# Ordered (substring -> category) rules; first match wins. Substrings are matched
# against the upper-cased Acyklama text.
_CATEGORY_RULES: list[tuple[str, str]] = [
    ('GUMRUKLEME', CustomsExpenseCategory.GUMRUKLEME),
    ('KARANTIN', CustomsExpenseCategory.KARANTIN),
    ('CT-1', CustomsExpenseCategory.CT1),
    ('CT -1', CustomsExpenseCategory.CT1),
    ('CT1', CustomsExpenseCategory.CT1),
    ('FITO', CustomsExpenseCategory.FITO),
    ('ANALIZ', CustomsExpenseCategory.ANALIZ),
    ('PASPORT SDELKA', CustomsExpenseCategory.PASPORT_SDELKA),
    ('PLATYOSKA', CustomsExpenseCategory.PLATYOSKA),
    ('POCTA', CustomsExpenseCategory.DOC_POST),
    ('YUZLENME', CustomsExpenseCategory.YUZLENME_HAT),
    ('GUMRUK AMAL', CustomsExpenseCategory.GUMRUK_AMAL),
    ('GRANITSA', CustomsExpenseCategory.BORDER_RETURN),
    ('YZYNA GAYDAN', CustomsExpenseCategory.BORDER_RETURN),
    ('SERTNAMA', CustomsExpenseCategory.SERTNAMA),
]

_QTY_RE = re.compile(r'(\d+)\s*(?:AD|SANY)', re.IGNORECASE)


def _norm_plate(value: object) -> str:
    """Strip every non-alphanumeric char and upper-case — for plate comparison."""
    return re.sub(r'[^A-Z0-9]', '', str(value).upper()) if value else ''


def _classify(desc: str) -> str:
    """Map an Acyklama description to a CustomsExpenseCategory code."""
    upper = desc.upper()
    for needle, category in _CATEGORY_RULES:
        if needle in upper:
            return category
    return CustomsExpenseCategory.OTHER


def _extract_quantity(desc: str) -> int | None:
    """Pull a unit count from a batch-fee label ('19 AD KARANTIN' -> 19)."""
    match = _QTY_RE.search(desc)
    return int(match.group(1)) if match else None


def _clean(value: object, limit: int) -> str | None:
    """Trim a cell to text, capped at ``limit`` chars; None when empty."""
    if value is None:
        return None
    text = str(value).strip()
    return text[:limit] if text else None


class Command(BaseCommand):
    help = 'Import the customs/document cash-advance ledger from data/avans.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=str(DEFAULT_PATH))
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Parse and report without writing to the database.',
        )
        parser.add_argument(
            '--user', type=str, default=None,
            help='Username to record as created_by/issued_by '
                 '(default: first superuser).',
        )

    def handle(self, *args, **options):
        path = Path(options['file'])
        dry_run = options['dry_run']
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        user = self._resolve_user(options['user'])
        plate_index = self._build_plate_index()

        advances, expenses, stats = self._parse(path, user, plate_index)

        self.stdout.write(
            f'Parsed {len(advances)} advances (sum={stats["advances_total"]} {CURRENCY}) '
            f'and {len(expenses)} expenses (sum={stats["expenses_total"]} {CURRENCY}).'
        )
        self.stdout.write(
            f'  Shipment links: {stats["linked"]} matched, '
            f'{stats["ambiguous"]} ambiguous (left null), '
            f'{stats["unmatched"]} unmatched/no-plate (left null).'
        )
        self.stdout.write('  By category: ' + ', '.join(
            f'{cat}={n}' for cat, n in sorted(stats['by_category'].items())
        ))

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — nothing written.'))
            return

        with transaction.atomic():
            removed = self._purge_previous()
            FinansistAdvance.objects.bulk_create(advances, batch_size=500)
            CustomsExpense.objects.bulk_create(expenses, batch_size=500)

        self.stdout.write(self.style.SUCCESS(
            f'Imported {len(advances)} advances + {len(expenses)} expenses '
            f'(removed {removed} prior import rows).'
        ))

    def _resolve_user(self, username: str | None) -> User:
        if username:
            try:
                return User.objects.get(username=username)
            except User.DoesNotExist as exc:
                raise CommandError(f'User not found: {username}') from exc
        user = User.objects.filter(is_superuser=True).order_by('id').first()
        if not user:
            raise CommandError('No superuser found; pass --user explicitly.')
        return user

    def _build_plate_index(self) -> dict[str, list[int]]:
        """Map normalized full truck_plate -> [shipment ids]."""
        index: dict[str, list[int]] = {}
        rows = (
            Shipment.objects
            .exclude(truck_plate__isnull=True)
            .exclude(truck_plate='')
            .values_list('id', 'truck_plate')
        )
        for sid, plate in rows:
            index.setdefault(_norm_plate(plate), []).append(sid)
        return index

    def _parse(self, path, user, plate_index):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[SHEET_NAME]

        advances: list[FinansistAdvance] = []
        expenses: list[CustomsExpense] = []
        stats = {
            'advances_total': Decimal('0'),
            'expenses_total': Decimal('0'),
            'linked': 0,
            'ambiguous': 0,
            'unmatched': 0,
            'by_category': {},
        }

        last_date: datetime.date | None = None
        for r in range(2, ws.max_row + 1):
            raw_date = ws.cell(r, COL_DATE).value
            if isinstance(raw_date, datetime.datetime):
                last_date = raw_date.date()

            desc = (ws.cell(r, COL_DESC).value or '').strip()
            amount_in = ws.cell(r, COL_IN).value
            amount_out = ws.cell(r, COL_OUT).value

            # A money row may omit its date when it continues the day above
            # (e.g. the "YER SANY YALNYS" correction fees) — carry it forward.
            has_amount = (
                (isinstance(amount_in, (int, float)) and amount_in)
                or (isinstance(amount_out, (int, float)) and amount_out)
            )
            if not has_amount or last_date is None:
                continue
            row_date = last_date

            if isinstance(amount_in, (int, float)) and amount_in:
                total = Decimal(str(amount_in))
                stats['advances_total'] += total
                advances.append(FinansistAdvance(
                    advance_date=row_date,
                    total_amount=total,
                    currency=CURRENCY,
                    purpose=_clean(desc, 200),
                    issued_by=user,
                    notes=IMPORT_MARKER,
                ))

            if isinstance(amount_out, (int, float)) and amount_out:
                total = Decimal(str(amount_out))
                stats['expenses_total'] += total
                category = _classify(desc)
                stats['by_category'][category] = stats['by_category'].get(category, 0) + 1

                shipment_id = self._match_shipment(
                    ws.cell(r, COL_PLATE).value, plate_index, stats
                )
                expenses.append(CustomsExpense(
                    expense_date=row_date,
                    category=category,
                    amount=total,
                    currency=CURRENCY,
                    shipment_id=shipment_id,
                    export_code_raw=_clean(ws.cell(r, COL_CODE).value, 50),
                    vehicle_plate=_clean(ws.cell(r, COL_PLATE).value, 60),
                    route_label=_clean(ws.cell(r, COL_FIRM).value, 120),
                    label_raw=_clean(desc, 255),
                    quantity=_extract_quantity(desc),
                    notes=IMPORT_MARKER,
                    created_by=user,
                ))

        return advances, expenses, stats

    def _match_shipment(self, plate, plate_index, stats) -> int | None:
        """Return a shipment id only on a UNIQUE full-plate match, else None."""
        if not plate:
            stats['unmatched'] += 1
            return None
        ids = set(plate_index.get(_norm_plate(plate), []))
        if len(ids) == 1:
            stats['linked'] += 1
            return next(iter(ids))
        if len(ids) > 1:
            stats['ambiguous'] += 1
        else:
            stats['unmatched'] += 1
        return None

    def _purge_previous(self) -> int:
        """Delete rows from a prior avans import (idempotent re-run)."""
        exp_deleted, _ = CustomsExpense.objects.filter(
            notes__contains=IMPORT_MARKER
        ).delete()
        adv_deleted, _ = FinansistAdvance.objects.filter(
            notes__contains=IMPORT_MARKER
        ).delete()
        return exp_deleted + adv_deleted
