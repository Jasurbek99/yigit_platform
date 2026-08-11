"""Import weekly harvest plans + truck allocations from weekly_plan.xlsx.

Source: data/weekly_plan.xlsx -> sheet 'Hepdelik planlama'

Structure: 29 week blocks (weeks 40-52/2025, weeks 1-16/2026).
Each block has:
  - 15 greenhouse rows with daily plan (Mon-Sat) and weekly actual total
  - Summary rows: Jemi (KG), Jemi Masyn Sany, Rossiya/Gazak/Gapy Satys Masyn Sany

Imports into:
  1. WeeklyHarvestPlan — one header row per (season, block, week, year)
  2. HarvestDayEntry — one row per (plan, day) carrying plan_value (cols C-H)
  3. WeeklyTruckAllocation — daily total_planned_kg from Jemi (KG) row
  4. TruckDestinationSplit — daily truck counts per destination (Rossiya/Gazak/Gapy Satys)

The weekly-actual column (col J) is NOT imported: the schema stores actuals
per-day (HarvestDayEntry.actual_value, sourced from shipment rollup), and a
weekly total cannot be split back into days.

Fill-empties semantics: nothing is deleted. Plans, day-entries, truck
allocations, and splits are created if missing. An existing plan cell is filled
only when its plan_value is NULL — operator-entered forecasts, actuals, and
daily-board data are never touched, and a plan_value already set is left as-is.
Re-running is therefore idempotent.
"""
import datetime
import logging
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).resolve().parents[5] / 'data' / 'weekly_plan.xlsx'
SHEET_NAME = 'Hepdelik maşyn sany'

BLOCK_NAME_PATTERNS = {
    'A-Ýyladyşhana': 'A',
    'B-Ýyladyşhana': 'B',
    'C-Ýyladyşhana': 'C',
    'D-Ýyladyşhana': 'D',
    'E-Ýyladyşhana': 'E',
    'F-Ýyladyşhana': 'F',
    'G-Ýyladyşhana': 'G',
    'H-Ýyladyşhana': 'H',
    'I-Ýyladyşhana': 'I',
    'J-Ýyladyşhana': 'J',
    'K-Ýyladyşhana': 'K',
    'L-Ýyladyşhana': 'L',
    'M15-Ýyladyşhana': 'M15',
    'M5-Ýyladyşhana': 'M5',
    'O-Ýyladyşhana': 'O',
}

# Row labels for truck summary data
TRUCK_ROW_LABELS = {
    'Jemi  (KG)': 'jemi_kg',
    'Jemi (KG)': 'jemi_kg',
    'Rossiya Masyn Sany': 'rossiya',
    'Gazak Masyn Sany': 'gazak',
    'Gapy Satys Masyn Sany': 'gapy_satys',
}

# Destination name -> label mapping (matches TruckDestination.name in DB)
DEST_DB_NAMES = {
    'rossiya': 'Rossiya',
    'gazak': 'Gazagystan',
    'gapy_satys': 'Gapy Satys',
}

SKIP_ROW_NAMES = {
    'Jemi Masyn Sany', 'Yyladyshanalar', 'Jogapkar',
}


def _parse_week_number(header: str) -> int | None:
    """Parse week number from strings like '40-NJY HEPDE' -> 40."""
    m = re.match(r'(\d+)', header.strip())
    return int(m.group(1)) if m else None


def _parse_kg_value(val) -> Decimal:
    """Parse a kg value that may be numeric, string-formatted, or None.

    Handles European-style thousands separators:
      '40,000,00' -> 40000.00
      '50.000.00' -> 50000.00

    Always quantizes to 2 decimal places (MSSQL DECIMAL(10,2)).
    """
    TWO_PLACES = Decimal('0.01')

    if val is None:
        return Decimal('0')
    if isinstance(val, (int, float)):
        d = Decimal(str(val)).quantize(TWO_PLACES)
        return d if d >= 0 else Decimal('0')

    s = str(val).strip().replace(' ', '')
    if not s or s == '-':
        return Decimal('0')

    s = s.replace(',', '.')
    parts = s.split('.')
    if len(parts) >= 3:
        integer_part = ''.join(parts[:-1])
        decimal_part = parts[-1]
        s = f'{integer_part}.{decimal_part}'

    try:
        d = Decimal(s).quantize(TWO_PLACES)
        return d if d >= 0 else Decimal('0')
    except InvalidOperation:
        logger.warning('Could not parse kg value: %r', val)
        return Decimal('0')


def _parse_truck_count(val) -> int:
    """Parse an integer truck count. Strings like 'bayramcylyk' -> 0."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return max(0, int(val))
    s = str(val).strip()
    try:
        return max(0, int(float(s)))
    except (ValueError, TypeError):
        return 0


def _match_block_code(cell_value: str) -> str | None:
    """Match a cell value to a GreenhouseBlock code."""
    for pattern, code in BLOCK_NAME_PATTERNS.items():
        if cell_value == pattern or cell_value.startswith(code + '-'):
            return code
    return None


class Command(BaseCommand):
    help = 'Import weekly harvest plans + truck allocations from weekly_plan.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?', default=str(DEFAULT_PATH))
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        from apps.core.models import GreenhouseBlock, TruckDestination
        from apps.core.seasons import get_active_season
        from apps.export.models import WeeklyTruckAllocation, TruckDestinationSplit
        from apps.greenhouse.models import HarvestDayEntry, WeeklyHarvestPlan

        path = Path(options['file'])
        if not path.exists():
            self.stderr.write(f'File not found: {path}')
            return

        dry_run = options['dry_run']

        season = get_active_season()
        if not season:
            self.stderr.write('No active season found.')
            return
        self.stdout.write(f'Season: {season.name} (id={season.id})')

        block_map = {b.code: b for b in GreenhouseBlock.objects.all()}
        self.stdout.write(f'Blocks: {sorted(block_map.keys())}')

        # Load TruckDestination map
        dest_map = {td.name: td for td in TruckDestination.objects.all()}
        self.stdout.write(f'Destinations: {list(dest_map.keys())}')

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            self.stderr.write(f'Sheet "{SHEET_NAME}" not found.')
            wb.close()
            return

        ws = wb[SHEET_NAME]

        # --- First pass: collect all data ---
        plan_entries = []
        # truck_data[week_number][year] = {
        #   'jemi_kg': [Mon, Tue, Wed, Thu, Fri, Sat],
        #   'rossiya': [Mon, ..., Sat],
        #   'gazak': [...], 'gapy_satys': [...]
        # }
        truck_data = {}
        skipped = 0
        warnings = []

        current_week_number = None
        current_year = None
        current_week_dates = []  # [Mon, Tue, Wed, Thu, Fri, Sat] dates or None
        in_week_block = False

        for row in ws.iter_rows(min_row=6, values_only=True):
            pad = list(row)
            while len(pad) < 11:
                pad.append(None)

            col0 = str(pad[0]).strip() if pad[0] else ''
            col1 = str(pad[1]).strip() if pad[1] else ''

            # --- Detect week header ---
            if 'HEPDE' in col0.upper():
                week_num = _parse_week_number(col0)
                if week_num is not None:
                    current_week_number = week_num
                    in_week_block = True
                    current_year = None
                    current_week_dates = []
                continue

            # --- Detect header row with dates ---
            if in_week_block and col1 == 'Yyladyshanalar':
                current_week_dates = []
                for di in range(2, 8):
                    dval = pad[di]
                    if isinstance(dval, datetime.datetime):
                        current_week_dates.append(dval.date())
                        if current_year is None:
                            current_year = dval.year
                    else:
                        current_week_dates.append(None)
                continue

            # --- Skip non-data rows ---
            if col1 in SKIP_ROW_NAMES or col0 in SKIP_ROW_NAMES:
                continue

            if not in_week_block or current_week_number is None:
                continue

            # --- Check if this is a truck summary row ---
            truck_label = TRUCK_ROW_LABELS.get(col1)
            if truck_label and current_year:
                week_key = (current_week_number, current_year)
                if week_key not in truck_data:
                    truck_data[week_key] = {}

                if truck_label == 'jemi_kg':
                    # Daily total kg (cols C-H)
                    truck_data[week_key]['jemi_kg'] = [
                        _parse_kg_value(pad[i]) for i in range(2, 8)
                    ]
                else:
                    # Truck counts per destination (cols C-H), integer
                    truck_data[week_key][truck_label] = [
                        _parse_truck_count(pad[i]) for i in range(2, 8)
                    ]
                continue

            # --- Match greenhouse block name ---
            block_code = _match_block_code(col1)
            if block_code is None:
                continue

            block = block_map.get(block_code)
            if block is None:
                warnings.append(f'W{current_week_number}: block {block_code!r} not in DB')
                skipped += 1
                continue

            if current_year is None:
                warnings.append(f'W{current_week_number}: no year — skipped {block_code}')
                skipped += 1
                continue

            if not current_week_dates or all(d is None for d in current_week_dates):
                warnings.append(f'W{current_week_number}: no dates in header — skipped {block_code}')
                skipped += 1
                continue

            plan_vals = [_parse_kg_value(pad[i]) for i in range(2, 8)]

            if all(v == Decimal('0') for v in plan_vals):
                skipped += 1
                continue

            plan_entries.append({
                'block': block,
                'week_dates': list(current_week_dates),
                'plan_vals': plan_vals,
            })

        wb.close()

        for w in warnings:
            self.stderr.write(f'WARNING: {w}')

        # --- Build truck allocation entries ---
        TRUCK_WEIGHT = Decimal('18500')
        TWO_PLACES = Decimal('0.01')
        DAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        truck_alloc_count = 0
        truck_split_count = 0

        # Count HarvestDayEntry rows that would be created (days with a plan > 0).
        day_entry_count = sum(
            1
            for rec in plan_entries
            for day_idx, val in enumerate(rec['plan_vals'])
            if rec['week_dates'][day_idx] is not None and val > Decimal('0')
        )

        if dry_run:
            # Count file-side candidates (truck allocs/splits with data this day)
            for (wn, yr), data in truck_data.items():
                jemi = data.get('jemi_kg', [Decimal('0')] * 6)
                for day_idx in range(6):
                    kg = jemi[day_idx]
                    if kg > 0:
                        truck_alloc_count += 1
                    for dest_label in ('rossiya', 'gazak', 'gapy_satys'):
                        counts = data.get(dest_label, [0] * 6)
                        if counts[day_idx] > 0:
                            truck_split_count += 1

            self.stdout.write(
                f'[dry-run] Fill-empties import for season {season.name} (no deletes).\n'
                f'Candidates from file (existing non-empty values are preserved):\n'
                f'  WeeklyHarvestPlan: {len(plan_entries)} block-weeks ({skipped} skipped)\n'
                f'  HarvestDayEntry plan cells: {day_entry_count}\n'
                f'  WeeklyTruckAllocation day-rows: ~{truck_alloc_count}\n'
                f'  TruckDestinationSplit: ~{truck_split_count}\n'
                f'  Warnings: {len(warnings)}'
            )
            return

        # === Fill-empties write: never delete, never overwrite non-empty values ===
        plan_created = 0
        day_created = 0
        day_filled = 0
        day_skipped = 0
        truck_alloc_created = 0
        truck_alloc_filled = 0
        truck_split_count = 0

        with transaction.atomic():
            # --- Harvest plans + per-day plan cells ---
            # ISO week/year are derived from each week's Monday so imported plans
            # share the same (season, block, week, year) key daily_board uses —
            # they dedupe with on-demand plans, not duplicate.
            for rec in plan_entries:
                block = rec['block']
                week_dates = rec['week_dates']
                plan_vals = rec['plan_vals']

                ref_date = next((d for d in week_dates if d is not None), None)
                if ref_date is None:
                    continue
                iso_year, iso_week, _ = ref_date.isocalendar()

                plan, plan_was_created = WeeklyHarvestPlan.objects.get_or_create(
                    season=season,
                    block=block,
                    week_number=iso_week,
                    year=iso_year,
                    defaults={'entered_by': None},
                )
                if plan_was_created:
                    plan_created += 1

                for day_idx, val in enumerate(plan_vals):
                    entry_date = week_dates[day_idx]
                    if entry_date is None or val <= Decimal('0'):
                        continue

                    entry, day_was_created = HarvestDayEntry.objects.get_or_create(
                        weekly_plan=plan,
                        entry_date=entry_date,
                        defaults={
                            'season': season,
                            'block': block,
                            'weekday': entry_date.weekday(),
                            'plan_value': val,
                        },
                    )
                    if day_was_created:
                        day_created += 1
                    elif entry.plan_value is None:
                        # Fill the empty plan cell; leave forecast/actual/daily intact.
                        entry.plan_value = val
                        entry.save(update_fields=['plan_value', 'updated_at'])
                        day_filled += 1
                    else:
                        day_skipped += 1

            # --- Truck allocations + destination splits ---
            for (wn, yr), data in sorted(truck_data.items()):
                jemi = data.get('jemi_kg', [Decimal('0')] * 6)

                for day_idx in range(6):
                    day_of_week = day_idx + 1  # 1=Monday .. 6=Saturday
                    kg = jemi[day_idx]

                    has_trucks = any(
                        data.get(dest_label, [0] * 6)[day_idx] > 0
                        for dest_label in ('rossiya', 'gazak', 'gapy_satys')
                    )
                    if kg == Decimal('0') and not has_trucks:
                        continue

                    trucks_calc = (kg / TRUCK_WEIGHT).quantize(TWO_PLACES) if kg > 0 else None

                    alloc, alloc_was_created = WeeklyTruckAllocation.objects.get_or_create(
                        season=season,
                        week_number=wn,
                        year=yr,
                        day_of_week=day_of_week,
                        defaults={
                            'total_planned_kg': kg if kg > 0 else None,
                            'total_trucks_calc': trucks_calc,
                            'decided_by': None,
                        },
                    )
                    if alloc_was_created:
                        truck_alloc_created += 1
                    else:
                        # Fill empty totals only; never overwrite existing decisions.
                        upd = []
                        if alloc.total_planned_kg is None and kg > 0:
                            alloc.total_planned_kg = kg
                            upd.append('total_planned_kg')
                        if alloc.total_trucks_calc is None and trucks_calc is not None:
                            alloc.total_trucks_calc = trucks_calc
                            upd.append('total_trucks_calc')
                        if upd:
                            alloc.save(update_fields=upd)
                            truck_alloc_filled += 1

                    # Destination splits: create only the ones not already present.
                    for dest_label in ('rossiya', 'gazak', 'gapy_satys'):
                        count = data.get(dest_label, [0] * 6)[day_idx]
                        if count <= 0:
                            continue
                        dest = dest_map.get(DEST_DB_NAMES[dest_label])
                        if not dest:
                            continue
                        _, split_was_created = TruckDestinationSplit.objects.get_or_create(
                            truck_allocation=alloc,
                            destination=dest,
                            defaults={'truck_count': count},
                        )
                        if split_was_created:
                            truck_split_count += 1

        self.stdout.write(self.style.SUCCESS(
            f'Fill-empties import complete ({skipped} block-weeks skipped):\n'
            f'  WeeklyHarvestPlan: {plan_created} new\n'
            f'  HarvestDayEntry plan cells: {day_created} new, {day_filled} filled, '
            f'{day_skipped} left untouched\n'
            f'  WeeklyTruckAllocation: {truck_alloc_created} new, {truck_alloc_filled} filled\n'
            f'  TruckDestinationSplit: {truck_split_count} new\n'
            f'  Warnings: {len(warnings)}'
        ))
