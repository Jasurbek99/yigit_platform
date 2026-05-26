"""Import daily domestic (local-market) sales from quota.xlsx into WeeklyLocalSellPlan.

Reads sheet "Kwota ucin icerki bazara berlen" — the kg each firm submitted to the
domestic market that earns export quota. Layout: firm rows 4-18 (col A name,
col B "Kabul edilen KG" total — skipped), one column per sale date (row 3),
trailing "Jemi tabsyrlan KG" total column — skipped.

WeeklyLocalSellPlan stores Mon-Sat per ISO week, so each daily value is folded
into its ISO week's day column. The lone Sunday date (2026-03-22, ISO 2026-W12)
is folded into that week's Saturday column.

Firm names use initials on this sheet ("Tel ED"); the shared FIRM_NAME_MAP maps
them to the spelled-out firms. NOTE: this fixes two mislabels from the earlier
import — "Tel G Amangeldiyew" → Tel Amangeldiyew G (not Tel Guwanc A.) and
"Tel GJ" → Tel Gurban J (not Tel Jumamyradow G).

Idempotent and non-destructive: update_or_create per (export_firm, week, year).
Existing rows for firms/weeks NOT in this file are left untouched. Updates
overwrite the Mon-Sat kg from the file but preserve the existing approval status;
new rows are created as 'approved'.

Usage:
    python manage.py import_local_sales                  # dry-run (default)
    python manage.py import_local_sales --commit         # write to DB
    python manage.py import_local_sales /path/to/file    # custom path
"""
import datetime
import logging
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.core.models import Season
from apps.export.models import WeeklyLocalSellPlan

from ._quota_import_utils import resolve_firm

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).parents[5] / 'data' / 'quota' / 'quota.xlsx'
SHEET_SALES = 'Kwota ucin icerki bazara berlen'

DATE_HEADER_ROW = 3
FIRST_FIRM_ROW = 4
# Total/derived rows in col A that are not firms.
NON_FIRM_LABELS = {'jemi', 'ugradylan kg', 'ara tapawut'}

DAY_COLS = (
    'monday_plan_kg', 'tuesday_plan_kg', 'wednesday_plan_kg',
    'thursday_plan_kg', 'friday_plan_kg', 'saturday_plan_kg',
)


class Command(BaseCommand):
    help = 'Import daily domestic sales from quota.xlsx into WeeklyLocalSellPlan'

    def add_arguments(self, parser):
        parser.add_argument('path', nargs='?', default=str(DEFAULT_PATH))
        parser.add_argument('--commit', action='store_true', help='Write to DB (default: dry-run)')

    def handle(self, *args, **options):
        path = Path(options['path'])
        commit = options['commit']

        if not path.exists():
            self.stderr.write(f'File not found: {path}')
            return

        season = Season.objects.filter(is_active=True).first()
        self.stdout.write(f'Loading {path} ...')
        self.stdout.write(f'  Active season: {season.name if season else "(none)"}')
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[SHEET_SALES]
        firm_cache: dict = {}

        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

        # Date columns: any column whose row-3 cell is a real date. This excludes
        # col B ("Kabul edilen KG") and the trailing "Jemi tabsyrlan KG" total.
        date_header = rows[DATE_HEADER_ROW - 1]
        date_by_idx: dict[int, datetime.date] = {
            idx: val.date()
            for idx, val in enumerate(date_header)
            if isinstance(val, datetime.datetime)
        }
        self.stdout.write(
            f'  Date columns: {len(date_by_idx)} '
            f'({min(date_by_idx.values())} -> {max(date_by_idx.values())})'
        )

        # accum[(firm_id, iso_year, iso_week)] = [mon..sat]
        accum: dict[tuple[int, int, int], list[Decimal]] = defaultdict(
            lambda: [Decimal('0')] * 6
        )
        sunday_folded = 0
        skipped_firms: list[str] = []
        firms_seen: set[int] = set()

        for row in rows[FIRST_FIRM_ROW - 1:]:
            name = row[0]
            if name is None or not str(name).strip():
                continue
            if str(name).strip().lower() in NON_FIRM_LABELS:
                continue
            firm = resolve_firm(str(name), firm_cache)
            if firm is None:
                skipped_firms.append(str(name).strip())
                continue
            firms_seen.add(firm.id)

            for idx, sale_date in date_by_idx.items():
                val = row[idx] if idx < len(row) else None
                if not val:
                    continue
                try:
                    kg = Decimal(str(val))
                except (ValueError, ArithmeticError):
                    continue
                if kg <= 0:
                    continue
                iso = sale_date.isocalendar()
                weekday = sale_date.weekday()   # 0=Mon .. 6=Sun
                if weekday == 6:                # Sunday → Saturday of same ISO week
                    sunday_folded += 1
                accum[(firm.id, iso[0], iso[1])][min(weekday, 5)] += kg

        self.stdout.write(f'\n  Firms resolved: {len(firms_seen)}')
        if skipped_firms:
            self.stdout.write(self.style.WARNING(
                f'  Firms NOT resolved (skipped): {", ".join(skipped_firms)}'
            ))
        if sunday_folded:
            self.stdout.write(f'  Sunday cells folded into Saturday: {sunday_folded}')
        self.stdout.write(f'  (firm, week, year) rows to upsert: {len(accum)}')
        grand = sum(sum(v) for v in accum.values())
        self.stdout.write(f'  Total kg across all rows: {grand:,.0f}')

        # Classify against existing rows: create vs update, and flag non-approved
        # (e.g. hand-entered draft) rows whose values an update would overwrite.
        firm_ids = {k[0] for k in accum}
        existing = {
            (r['export_firm_id'], r['year'], r['week_number']): r['status']
            for r in WeeklyLocalSellPlan.objects.filter(
                export_firm_id__in=firm_ids,
            ).values('export_firm_id', 'year', 'week_number', 'status')
        }
        to_create = [k for k in accum if k not in existing]
        to_update = [k for k in accum if k in existing]
        draft_overwrites = [k for k in to_update if existing[k] != 'approved']
        self.stdout.write(f'  -> would CREATE {len(to_create)}, UPDATE {len(to_update)}')
        if draft_overwrites:
            self.stdout.write(self.style.WARNING(
                f'  -> {len(draft_overwrites)} update(s) currently non-approved; their Mon-Sat '
                f'values would be overwritten (existing status preserved).'
            ))

        if not commit:
            self.stdout.write('\n  DRY RUN — use --commit to write to DB')
            return

        with transaction.atomic():
            created = updated = 0
            for (firm_id, year, week), day_vals in accum.items():
                defaults = dict(zip(DAY_COLS, day_vals))
                defaults['season'] = season
                obj, was_created = WeeklyLocalSellPlan.objects.update_or_create(
                    export_firm_id=firm_id, year=year, week_number=week,
                    defaults=defaults,
                )
                if was_created:
                    obj.status = 'approved'   # historical actuals; preserve status on update
                    obj.save(update_fields=['status'])
                    created += 1
                else:
                    updated += 1

            self.stdout.write(self.style.SUCCESS(
                f'\n  Done: {created} created, {updated} updated.'
            ))
