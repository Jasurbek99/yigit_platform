"""Import domestic market prices from Satys_bahalar_202526.xlsx → DomesticMarketPrice.

Source: Satys_bahalar_202526.xlsx — 7 monthly sheets (Sep 2025–Mar 2026)

Each sheet has three sections per row (three panels side by side):
  BAZAR section:   cols 0-6  (date, market, Salkym, Gulpakly, Gulpaksyz, Mayda, Cherri)
  KLENTLER section: cols 8-14 (date, market, Salkym, Gulpakly, Gulpaksyz, Mayda, Cherri)
  Onlayn section:   cols 16-22 (date, market, Salkym, Gulpakly, Gulpaksyz, Gulgune, Cherry)

Each combination of (date, market_name, price_type, variety_type) → one DomesticMarketPrice row.

price_type values: 'bazar', 'klent', 'online'
variety_type values: 'tomato_salkym', 'tomato_gulpakly', 'tomato_gulpaksyz', 'tomato_mayda',
                     'tomato_cherri', 'tomato_gulgune'

Skip rows:
  - Aggregate rows: Boluleni, Klient sanyna, Ortaca, Ortalyk, Real, Rashot, BAZAR, KLENTLER header
  - Cells where price is None or not a number
  - Market names that are dates (datetime objects in market column)
"""
import datetime
import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).parents[6] / 'data' / 'p3-export' / 'Satys_bahalar_202526.xlsx'

# Sheets to process (in order)
MONTHLY_SHEETS = [
    'Sentyabr 25',
    'Oktyabr 25',
    'Noyabr 25',
    'Dekabr 25',
    'Yanwar 26',
    'Fewral 26',
    'Mart 26',
]

# Rows to skip (aggregate/header rows in market name column)
SKIP_MARKETS = {
    'Boluleni', 'Klient sanyna', 'Ortaca:', 'Ortalyk', 'Real:', 'Rashot',
    'BAZAR', 'KLENTLER', 'Onlayn bahalar', 'Jemi', '',
}

# Column layout for each section panel:
# (date_col, market_col, variety_cols_with_names)
PANELS = [
    # BAZAR
    ('bazar', 0, 1, [
        (2, 'tomato_salkym'),
        (3, 'tomato_gulpakly'),
        (4, 'tomato_gulpaksyz'),
        (5, 'tomato_mayda'),
        (6, 'tomato_cherri'),
    ]),
    # KLENTLER
    ('klent', 8, 9, [
        (10, 'tomato_salkym'),
        (11, 'tomato_gulpakly'),
        (12, 'tomato_gulpaksyz'),
        (13, 'tomato_mayda'),
        (14, 'tomato_cherri'),
    ]),
    # Onlayn
    ('online', 16, 17, [
        (18, 'tomato_salkym'),
        (19, 'tomato_gulpakly'),
        (20, 'tomato_gulpaksyz'),
        (21, 'tomato_gulgune'),
        (22, 'tomato_cherri'),
    ]),
]


def _to_decimal(val):
    if val is None:
        return None
    try:
        d = Decimal(str(val))
        if d <= 0:
            return None
        return d
    except (InvalidOperation, ValueError):
        return None


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


class Command(BaseCommand):
    help = 'Import domestic market prices from Satys_bahalar_202526.xlsx → DomesticMarketPrice'

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?', default=str(DEFAULT_PATH))
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        from apps.export.models import DomesticMarketPrice

        path = Path(options['file'])
        if not path.exists():
            self.stderr.write(f'File not found: {path}')
            return

        dry_run = options['dry_run']

        entries = []
        skipped = 0
        warnings = []

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

        for sheet_name in MONTHLY_SHEETS:
            if sheet_name not in wb.sheetnames:
                self.stderr.write(f'WARNING: Sheet "{sheet_name}" not found — skipping')
                continue

            ws = wb[sheet_name]
            sheet_entries = 0

            # Track last seen date per panel (dates may repeat across rows on same date)
            last_date = [None, None, None]  # indexed by panel position

            for row in ws.iter_rows(min_row=3, values_only=True):
                pad = list(row)
                while len(pad) < 23:
                    pad.append(None)

                for panel_idx, (price_type, date_col, market_col, variety_cols) in enumerate(PANELS):
                    raw_date = pad[date_col]
                    raw_market = pad[market_col]

                    # Update last known date if this cell has a date
                    d = _parse_date(raw_date)
                    if d is not None:
                        last_date[panel_idx] = d

                    use_date = last_date[panel_idx]
                    if use_date is None:
                        skipped += 1
                        continue

                    # Get market name
                    if raw_market is None or isinstance(raw_market, datetime.datetime):
                        skipped += 1
                        continue

                    market_name = str(raw_market).strip()
                    if not market_name or market_name in SKIP_MARKETS:
                        skipped += 1
                        continue

                    # Emit one DomesticMarketPrice per (date, market, price_type, variety) cell
                    for (col_idx, variety_type) in variety_cols:
                        price_val = pad[col_idx] if col_idx < len(pad) else None
                        price = _to_decimal(price_val)
                        if price is None:
                            continue

                        entries.append(DomesticMarketPrice(
                            date=use_date,
                            market_name=market_name,
                            price_type=price_type,
                            variety_type=variety_type,
                            price=price,
                            entered_by=None,
                        ))
                        sheet_entries += 1

            self.stdout.write(f'  Sheet "{sheet_name}": {sheet_entries} entries parsed')

        wb.close()

        for w in warnings:
            self.stderr.write(f'WARNING: {w}')

        if dry_run:
            self.stdout.write(
                f'[dry-run] Would import {len(entries)} DomesticMarketPrice rows '
                f'({skipped} cells skipped) | Warnings: {len(warnings)}'
            )
            return

        created = 0
        with transaction.atomic():
            for i in range(0, len(entries), 500):
                batch = entries[i:i + 500]
                result = DomesticMarketPrice.objects.bulk_create(
                    batch, batch_size=500, ignore_conflicts=True
                )
                created += len(result)

        self.stdout.write(self.style.SUCCESS(
            f'Imported: {created} | Skipped: {skipped} | Warnings: {len(warnings)}'
        ))
