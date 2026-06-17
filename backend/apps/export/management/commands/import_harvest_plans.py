"""Import weekly harvest plans from Pomidor_Dükany__20252026.xlsx.

Source: Pomidor_Dükany__20252026.xlsx → sheet 'Hepdelik planlama'

Structure: The sheet is divided into week blocks, each starting with a 'XX-NJY HEPDE' header.
Each week block contains:
  - Header row: Jogapkar | Yyladyshanalar | Mon | Tue | Wed | Thu | Fri | Sat | Jemi | Actual
  - 15 data rows: one per greenhouse block (A-L, M15, M5, O)
  - Total/summary rows (Jemi, truck counts) — skipped

Each block row maps to:
  1. WeeklyHarvestPlan — one header per (season, block, week, year). ISO week/year
     are derived from each week's Monday so headers share the get_or_create key
     daily_board uses (imported plans dedupe with on-demand ones, not duplicate).
  2. HarvestDayEntry — one row per (plan, day) carrying plan_value (cols C-H).

The wide *_plan_kg columns and the weekly-actual column were dropped from the
schema (migration greenhouse.0004) — daily plan/actual data now lives in
HarvestDayEntry. The weekly-actual total (col J) is NOT imported: actuals are
stored per-day (sourced from shipment rollup) and a weekly total cannot be split
back into days.

Fill-empties semantics (idempotent, non-destructive): nothing is deleted. Plans
and day-entries are created if missing; an existing plan cell is filled only when
its plan_value is NULL — operator-entered forecasts, actuals, and daily-board
data are never touched, and a plan_value already set is left as-is.

Skip rows:
  - Block names not matching known GreenhouseBlock codes
  - Rows where ALL plan values are None or 0
  - Aggregate rows (Jemi, Masyn Sany, etc.)
"""
import datetime
import logging
import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).resolve().parents[5] / 'data' / 'p3-export' / 'Pomidor_Dükany__20252026.xlsx'
SHEET_NAME = 'Hepdelik planlama'

# Block name patterns in col 1 → GreenhouseBlock.code
BLOCK_NAME_PATTERNS = {
    'A-Ýyladyşhana': 'A',
    'B-Ýyladyşhana': 'B',
    'C-Ýyladyşhana': 'C',
    'D-Ýyladyşhana': 'D',
    'E-Ýyladyşhana': 'E',
    'F-Ýyladyşhana': 'F',
    'G-Ýyladyşhana': 'G',
    'H-Ýyladyşhana': 'H',
    'I-Ýyladyşhana': 'I',
    'J-Ýyladyşhana': 'J',
    'K-Ýyladyşhana': 'K',
    'L-Ýyladyşhana': 'L',
    'M15-Ýyladyşhana': 'M15',
    'M5-Ýyladyşhana': 'M5',
    'O-Ýyladyşhana': 'O',
}

# Rows to skip by name in col 1
SKIP_ROW_NAMES = {
    'Jemi  (KG)', 'Jemi Masyn Sany', 'Rossiya Masyn Sany', 'Gazak Masyn Sany',
    'Gapy Satys Masyn Sany', 'Yyladyshanalar', 'Jogapkar',
}


def _to_decimal_or_zero(val):
    """Convert numeric value to Decimal; return Decimal(0) if None/zero."""
    if val is None:
        return Decimal('0')
    try:
        d = Decimal(str(val))
        return d if d >= 0 else Decimal('0')
    except Exception:
        return Decimal('0')


def _parse_week_number(header: str):
    """Parse week number from strings like '40-NJY HEPDE' → 40."""
    m = re.match(r'(\d+)', header.strip())
    if m:
        return int(m.group(1))
    return None


class Command(BaseCommand):
    help = 'Import weekly harvest plans from Pomidor_Dükany__20252026.xlsx → WeeklyHarvestPlan'

    def add_arguments(self, parser):
        parser.add_argument('file', nargs='?', default=str(DEFAULT_PATH))
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        from apps.greenhouse.models import HarvestDayEntry, WeeklyHarvestPlan
        from apps.core.models import GreenhouseBlock, Season

        path = Path(options['file'])
        if not path.exists():
            self.stderr.write(f'File not found: {path}')
            return

        dry_run = options['dry_run']

        # Get active season
        season = Season.objects.filter(is_active=True).first()
        if not season:
            self.stderr.write('No active season found — cannot import harvest plans.')
            return
        self.stdout.write(f'Using season: {season.name} (id={season.id})')

        # Pre-load GreenhouseBlock cache
        block_map = {b.code: b for b in GreenhouseBlock.objects.all()}
        self.stdout.write(f'Loaded {len(block_map)} greenhouse blocks: {sorted(block_map.keys())}')

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            self.stderr.write(f'Sheet "{SHEET_NAME}" not found in workbook.')
            wb.close()
            return

        ws = wb[SHEET_NAME]

        entries = []
        skipped = 0
        warnings = []

        # State machine for parsing week blocks
        current_week_number = None
        current_year = None
        current_week_dates = []  # [Mon, Tue, Wed, Thu, Fri, Sat]
        in_week_block = False

        for row in ws.iter_rows(min_row=6, values_only=True):
            pad = list(row)
            while len(pad) < 11:
                pad.append(None)

            col0 = str(pad[0]).strip() if pad[0] else ''
            col1 = str(pad[1]).strip() if pad[1] else ''

            # --- Detect week header ---
            if 'HEPDE' in col0:
                week_num = _parse_week_number(col0)
                if week_num is not None:
                    current_week_number = week_num
                    in_week_block = True
                    current_week_dates = []
                    # Year determined from dates in header row (next non-empty row)
                    current_year = None
                continue

            # --- Detect header row with dates ---
            if in_week_block and col1 == 'Yyladyshanalar':
                # cols 2-7 are the week dates (Mon-Sat)
                current_week_dates = []
                for di in range(2, 8):
                    dval = pad[di]
                    if isinstance(dval, datetime.datetime):
                        current_week_dates.append(dval.date())
                        if current_year is None:
                            current_year = dval.year
                    else:
                        current_week_dates.append(None)
                continue

            # --- Skip aggregate/summary rows ---
            if col1 in SKIP_ROW_NAMES:
                continue

            if not in_week_block or current_week_number is None:
                continue

            # --- Try to match a block name ---
            block_code = None
            for pattern, code in BLOCK_NAME_PATTERNS.items():
                if col1.startswith(pattern.split('-')[0] + '-') or col1.strip() == pattern.strip():
                    block_code = code
                    break
            if block_code is None:
                # Try simpler prefix match
                for pattern, code in BLOCK_NAME_PATTERNS.items():
                    if col1.startswith(code + '-'):
                        block_code = code
                        break

            if block_code is None:
                continue

            block = block_map.get(block_code)
            if block is None:
                warnings.append(f'Week {current_week_number}: block code {block_code!r} not in DB — skipped')
                skipped += 1
                continue

            if current_year is None:
                warnings.append(f'Week {current_week_number}: no year detected — skipped block {block_code}')
                skipped += 1
                continue

            if not current_week_dates or all(d is None for d in current_week_dates):
                warnings.append(f'Week {current_week_number}: no dates in header — skipped block {block_code}')
                skipped += 1
                continue

            # Extract plan values (cols 2-7). The weekly-actual column (col 9) is
            # not imported — actuals are per-day (shipment rollup) and a weekly
            # total cannot be split back into days.
            plan_vals = [_to_decimal_or_zero(pad[i]) for i in range(2, 8)]

            # Skip rows where all plan values are 0
            if all(v == Decimal('0') for v in plan_vals):
                skipped += 1
                continue

            entries.append({
                'block': block,
                'week_dates': list(current_week_dates),
                'plan_vals': plan_vals,
            })

        wb.close()

        for w in warnings:
            self.stderr.write(f'WARNING: {w}')

        # Count HarvestDayEntry rows that would be created (days with a plan > 0).
        day_entry_count = sum(
            1
            for rec in entries
            for day_idx, val in enumerate(rec['plan_vals'])
            if rec['week_dates'][day_idx] is not None and val > Decimal('0')
        )

        if dry_run:
            self.stdout.write(
                f'[dry-run] Fill-empties import for season {season.name} (no deletes).\n'
                f'Candidates from file (existing non-empty values are preserved):\n'
                f'  WeeklyHarvestPlan: {len(entries)} block-weeks ({skipped} skipped)\n'
                f'  HarvestDayEntry plan cells: {day_entry_count}\n'
                f'  Warnings: {len(warnings)}'
            )
            return

        # === Fill-empties write: never delete, never overwrite non-empty values ===
        plan_created = 0
        day_created = 0
        day_filled = 0
        day_skipped = 0

        with transaction.atomic():
            for rec in entries:
                block = rec['block']
                week_dates = rec['week_dates']
                plan_vals = rec['plan_vals']

                # ISO week/year from the week's Monday so imported plans share the
                # (season, block, week, year) key daily_board uses.
                ref_date = next((d for d in week_dates if d is not None), None)
                if ref_date is None:
                    continue
                iso_year, iso_week, _ = ref_date.isocalendar()

                plan, plan_was_created = WeeklyHarvestPlan.objects.get_or_create(
                    season=season,
                    block=block,
                    week_number=iso_week,
                    year=iso_year,
                    defaults={'entered_by': None},
                )
                if plan_was_created:
                    plan_created += 1

                for day_idx, val in enumerate(plan_vals):
                    entry_date = week_dates[day_idx]
                    if entry_date is None or val <= Decimal('0'):
                        continue

                    entry, day_was_created = HarvestDayEntry.objects.get_or_create(
                        weekly_plan=plan,
                        entry_date=entry_date,
                        defaults={
                            'season': season,
                            'block': block,
                            'weekday': entry_date.weekday(),
                            'plan_value': val,
                        },
                    )
                    if day_was_created:
                        day_created += 1
                    elif entry.plan_value is None:
                        # Fill the empty plan cell; leave forecast/actual/daily intact.
                        entry.plan_value = val
                        entry.save(update_fields=['plan_value', 'updated_at'])
                        day_filled += 1
                    else:
                        day_skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f'Fill-empties import complete ({skipped} block-weeks skipped):\n'
            f'  WeeklyHarvestPlan: {plan_created} new\n'
            f'  HarvestDayEntry plan cells: {day_created} new, {day_filled} filled, '
            f'{day_skipped} left untouched\n'
            f'  Warnings: {len(warnings)}'
        ))
