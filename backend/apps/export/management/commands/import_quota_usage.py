"""Import quota usage (Islenen Kwota) from quota.xlsx into QuotaUsageRecord.

Reads sheet Kwota-2, "used quota" section:
  - Tomato: firm headers row 8 cols B-P (15 firms), date col A, data rows 33-108.
  - Pepper: firm headers row 8 cols Y-Z (2 firms), date col X, data rows 32-39.
Each non-zero cell becomes one kg-of-quota-consumed record per (date, firm, product).
Repeated dates (e.g. two 19.03.2026 rows) are aggregated.

Idempotent: deletes prior rows with notes='Imported from quota.xlsx' before reload.

Usage:
    python manage.py import_quota_usage                  # dry-run (default)
    python manage.py import_quota_usage --commit         # write to DB
    python manage.py import_quota_usage /path/to/file    # custom path
"""
import logging
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.export.models import QuotaUsageRecord

from ._quota_import_utils import parse_quota_date, resolve_firm

logger = logging.getLogger(__name__)

DEFAULT_PATH = Path(__file__).parents[5] / 'data' / 'quota' / 'quota.xlsx'
SHEET_QUOTA = 'Kwota-2'

HEADER_ROW = 8

# Tomato: 15 firm columns B(2)-P(16), date in col A(1), usage data rows 33-108.
TOMATO_FIRM_COLS = list(range(2, 17))
TOMATO_DATE_COL = 1
TOMATO_USAGE_ROWS = list(range(33, 109))

# Pepper: 2 firm columns Y(25)-Z(26), date in col X(24), usage data rows 32-39.
PEPPER_FIRM_COLS = [25, 26]
PEPPER_DATE_COL = 24
PEPPER_USAGE_ROWS = list(range(32, 40))


class Command(BaseCommand):
    help = 'Import used quota (Islenen Kwota) from quota.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('path', nargs='?', default=str(DEFAULT_PATH))
        parser.add_argument('--commit', action='store_true', help='Write to DB (default: dry-run)')

    def handle(self, *args, **options):
        path = Path(options['path'])
        commit = options['commit']

        if not path.exists():
            self.stderr.write(f'File not found: {path}')
            return

        self.stdout.write(f'Loading {path} ...')
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[SHEET_QUOTA]
        firm_cache: dict = {}

        # Aggregate same (date, product, firm) — duplicate date rows exist.
        agg: dict[tuple[date, str, str], Decimal] = defaultdict(Decimal)

        def _read_section(firm_cols, date_col, data_rows, product):
            col_firms = {
                col: str(ws.cell(row=HEADER_ROW, column=col).value).strip()
                for col in firm_cols
                if ws.cell(row=HEADER_ROW, column=col).value
            }
            self.stdout.write(
                f'  {product.title()} usage: {len(col_firms)} firms, '
                f'rows {data_rows[0]}-{data_rows[-1]}'
            )

            for row_idx in data_rows:
                usage_date = parse_quota_date(
                    ws.cell(row=row_idx, column=date_col).value, fix_out_of_season=True,
                )
                if not usage_date:
                    continue
                for col, firm_name in col_firms.items():
                    val = ws.cell(row=row_idx, column=col).value
                    if not val:
                        continue
                    try:
                        kg = Decimal(str(val))
                    except (ValueError, ArithmeticError):
                        continue
                    if kg <= 0:
                        continue
                    agg[(usage_date, product, firm_name)] += kg

        _read_section(TOMATO_FIRM_COLS, TOMATO_DATE_COL, TOMATO_USAGE_ROWS, 'tomato')
        _read_section(PEPPER_FIRM_COLS, PEPPER_DATE_COL, PEPPER_USAGE_ROWS, 'pepper')

        self.stdout.write(f'\n  Unique (date, product, firm) records: {len(agg)}')
        for product in ('tomato', 'pepper'):
            total = sum(v for (_, p, _), v in agg.items() if p == product)
            count = sum(1 for (_, p, _) in agg if p == product)
            self.stdout.write(f'  {product.title():6}: {count:3} records, {total:>12,.0f} kg total')

        # Confirm the idempotent delete will sweep exactly the prior import.
        existing_total = QuotaUsageRecord.objects.count()
        sweepable = QuotaUsageRecord.objects.filter(notes='Imported from quota.xlsx').count()
        self.stdout.write(
            f'\n  Existing QuotaUsageRecord rows: {existing_total} '
            f'({sweepable} match notes="Imported from quota.xlsx" and will be replaced)'
        )
        if existing_total != sweepable:
            self.stdout.write(self.style.WARNING(
                f'  {existing_total - sweepable} row(s) came from elsewhere and will NOT be deleted.'
            ))

        if not commit:
            self.stdout.write('\n  DRY RUN — use --commit to write to DB')
            return

        with transaction.atomic():
            deleted, _ = QuotaUsageRecord.objects.filter(
                notes='Imported from quota.xlsx',
            ).delete()
            if deleted:
                self.stdout.write(f'  Deleted {deleted} previously imported usage records')

            objs, skipped = [], 0
            for (usage_date, product, firm_name), kg in sorted(agg.items()):
                firm = resolve_firm(firm_name, firm_cache)
                if firm is None:
                    skipped += 1
                    continue
                objs.append(QuotaUsageRecord(
                    usage_date=usage_date,
                    export_firm=firm,
                    kg_used=kg,
                    product_type=product,
                    status='approved',
                    notes='Imported from quota.xlsx',
                ))
            QuotaUsageRecord.objects.bulk_create(objs, batch_size=500)

            self.stdout.write(self.style.SUCCESS(
                f'\n  Done: {len(objs)} usage records created, '
                f'{len(firm_cache)} firms resolved, {skipped} skipped (firm not found).'
            ))
